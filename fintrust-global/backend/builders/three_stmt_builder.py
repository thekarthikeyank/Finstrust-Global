"""
3-Statement Integrated Model Builder — 9 Sheets
IS → BS → CFS fully linked with Working Capital & Debt Schedule
"""

from builders.base_builder import BaseBuilder
from formatting.institutional import Colors, Fonts, Fills, Borders, NumFormats, ColWidths, Formatter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel


class ThreeStatementBuilder(BaseBuilder):

    def __init__(self, assumptions: dict):
        super().__init__(assumptions)
        self.a = assumptions
        self.years = int(self.a.get("projection_years", 5))
        self.base_year = int(self.a.get("base_year", 2024))
        self.dc = 3

    def build(self):
        self.build_cover("3-Statement Integrated")
        self._build_assumptions()
        self._build_income_statement()
        self._build_working_capital()
        self._build_debt_schedule()
        self._build_cash_flow_statement()
        self._build_balance_sheet()
        self._build_checks()
        self._build_dashboard()
        return self.save_to_buffer()

    def _yr_header(self, ws, row, n, dc):
        for i in range(n):
            col = dc + i
            cell = ws.cell(row=row, column=col, value=f"FY{self.base_year + i + 1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

    def _build_assumptions(self):
        ws, fmt = self.add_sheet("ASSUMPTIONS", tab_color="2E75B6")
        a = self.a
        fmt.set_column_widths({"A": 3, "B": 40, "C": 22})
        fmt.apply_sheet_title(2, 2, "MODEL ASSUMPTIONS", "3-Statement Integrated Model Drivers")
        fmt.apply_units_label(3, 2, "$ in Millions")

        sections = [
            (5, "REVENUE & MARGIN", [
                ("Base Year Revenue ($M)",     "base_revenue",  200.0, NumFormats.USD_MILLIONS),
                ("Annual Revenue Growth %",    "rev_growth",    0.08,  NumFormats.PERCENT_ONE),
                ("Gross Margin %",             "gross_margin",  0.55,  NumFormats.PERCENT_ONE),
                ("EBITDA Margin %",            "ebitda_margin", 0.22,  NumFormats.PERCENT_ONE),
                ("D&A as % Revenue",           "da_pct",        0.05,  NumFormats.PERCENT_ONE),
                ("Interest Rate on Debt",      "interest_rate", 0.06,  NumFormats.PERCENT_ONE),
                ("Tax Rate",                   "tax_rate",      0.25,  NumFormats.PERCENT_ONE),
                ("Dividend Payout Ratio",      "dividend_pct",  0.20,  NumFormats.PERCENT_ONE),
            ]),
            (15, "CAPEX & WORKING CAPITAL", [
                ("Capex as % Revenue",         "capex_pct",     0.06,  NumFormats.PERCENT_ONE),
                ("Days Sales Outstanding",     "dso",           45,    NumFormats.INTEGER),
                ("Days Inventory Outstanding", "dio",           60,    NumFormats.INTEGER),
                ("Days Payable Outstanding",   "dpo",           30,    NumFormats.INTEGER),
            ]),
            (21, "BALANCE SHEET — OPENING", [
                ("Opening Cash ($M)",          "open_cash",     20.0,  NumFormats.USD_MILLIONS),
                ("Opening Debt ($M)",          "open_debt",     50.0,  NumFormats.USD_MILLIONS),
                ("Opening PP&E ($M)",          "open_ppe",      80.0,  NumFormats.USD_MILLIONS),
                ("Opening Equity ($M)",        "open_equity",   100.0, NumFormats.USD_MILLIONS),
            ]),
        ]

        for header_row, section_label, items in sections:
            fmt.apply_header_row(header_row, 2, 4, section_label)
            for i, (lbl, key, default, fmt_code) in enumerate(items):
                fmt.apply_label_cell(header_row + 1 + i, 2, lbl)
                val = float(a.get(key, default))
                fmt.apply_input_cell(header_row + 1 + i, 3, val, fmt_code)

    def _build_income_statement(self):
        ws, fmt = self.add_sheet("INCOME_STATEMENT", tab_color="375623")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "INCOME STATEMENT", "Projected Profit & Loss")
        fmt.apply_units_label(3, 2, "$ in Millions")
        self._yr_header(ws, 4, n, dc)

        row_defs = [
            (5,  "REVENUE",                      True),
            (6,  "Net Revenue",                  False),
            (8,  "COST STRUCTURE",               True),
            (9,  "Cost of Goods Sold",            False),
            (10, "Gross Profit",                 False),
            (12, "OPERATING EXPENSES",           True),
            (13, "EBITDA",                       False),
            (14, "Depreciation & Amortization",  False),
            (15, "EBIT",                         False),
            (17, "BELOW THE LINE",               True),
            (18, "Interest Expense",             False),
            (19, "EBT",                          False),
            (20, "Income Tax",                   False),
            (21, "Net Income",                   False),
            (23, "EARNINGS DISTRIBUTION",        True),
            (24, "Dividends Paid",               False),
            (25, "Addition to Retained Earnings",False),
            (27, "MARGIN ANALYSIS",              True),
            (28, "Gross Margin %",               False),
            (29, "EBITDA Margin %",              False),
            (30, "Net Margin %",                 False),
            (31, "Revenue Growth %",             False),
        ]

        for row_num, label, is_hdr in row_defs:
            if is_hdr:
                fmt.apply_header_row(row_num, 2, dc + n - 1, label)
            else:
                fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            rev = f"={asm}!C6*(1+{asm}!C7)^{i+1}"
            formulas = {
                6:  rev,
                9:  f"=-{cl}6*(1-{asm}!C8)",
                10: f"={cl}6+{cl}9",
                13: f"={cl}6*{asm}!C9",
                14: f"=-{cl}6*{asm}!C10",
                15: f"={cl}13+{cl}14",
                18: f"=-DEBT_SCHEDULE!{cl}5*{asm}!C11",
                19: f"={cl}15+{cl}18",
                20: f"=MAX(-{cl}19*{asm}!C12,0)",
                21: f"={cl}19+{cl}20",
                24: f"=-{cl}21*{asm}!C13",
                25: f"={cl}21+{cl}24",
                28: f"={cl}10/{cl}6",
                29: f"={cl}13/{cl}6",
                30: f"={cl}21/{cl}6",
                31: f"=IF({i}>0,{cl}6/{prev_cl}6-1,\"\")",
            }
            pct_rows = [28, 29, 30, 31]
            total_rows = [10, 13, 15, 19, 21]

            for row_num, formula in formulas.items():
                c = ws.cell(row=row_num, column=col, value=formula)
                c.alignment = Alignment(horizontal="right")
                if row_num in total_rows:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                    c.number_format = NumFormats.USD_MILLIONS
                elif row_num in pct_rows:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()
                    c.number_format = NumFormats.PERCENT_ONE
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()
                    c.number_format = NumFormats.USD_MILLIONS

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_working_capital(self):
        ws, fmt = self.add_sheet("WORKING_CAPITAL", tab_color="7030A0")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"
        is_s = "INCOME_STATEMENT"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "WORKING CAPITAL SCHEDULE", "DSO / DIO / DPO → NWC Change")
        fmt.apply_units_label(3, 2, "$ in Millions")
        self._yr_header(ws, 4, n, dc)

        wc_rows = [
            (5,  "WORKING CAPITAL DRIVERS",   True),
            (6,  "Revenue",                    False),
            (7,  "COGS",                       False),
            (9,  "NWC COMPONENTS",             True),
            (10, "Accounts Receivable (DSO)",  False),
            (11, "Inventory (DIO)",            False),
            (12, "Accounts Payable (DPO)",     False),
            (13, "Net Working Capital",        False),
            (15, "NWC CHANGE",                 True),
            (16, "Change in NWC (↑ = use)",   False),
        ]

        for row_num, label, is_hdr in wc_rows:
            if is_hdr:
                fmt.apply_header_row(row_num, 2, dc + n - 1, label)
            else:
                fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            formulas = {
                6:  f"={is_s}!{cl}6",
                7:  f"=-{is_s}!{cl}9",
                10: f"={cl}6*{asm}!C16/365",
                11: f"={cl}7*{asm}!C17/365",
                12: f"=-{cl}7*{asm}!C18/365",
                13: f"={cl}10+{cl}11+{cl}12",
                16: f"=IF({i}>0,{cl}13-{prev_cl}13,{cl}13)",
            }

            for row_num, formula in formulas.items():
                c = ws.cell(row=row_num, column=col, value=formula)
                c.alignment = Alignment(horizontal="right")
                c.number_format = NumFormats.USD_MILLIONS
                if row_num in [13]:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_debt_schedule(self):
        ws, fmt = self.add_sheet("DEBT_SCHEDULE", tab_color="C00000")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "DEBT SCHEDULE", "Revolver & Term Loan Schedule")
        fmt.apply_units_label(3, 2, "$ in Millions")
        self._yr_header(ws, 4, n, dc)

        fmt.apply_header_row(5, 2, dc + n - 1, "TERM LOAN")
        for row_num, label in [(6, "Opening Balance"), (7, "Repayment (5% p.a.)"), (8, "Closing Balance")]:
            fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                open_b = f"={asm}!C22"
            else:
                open_b = f"={prev_cl}8"

            for row_num, formula in [
                (6, open_b),
                (7, f"=-{cl}6*0.05"),
                (8, f"=MAX({cl}6+{cl}7,0)"),
            ]:
                c = ws.cell(row=row_num, column=col, value=formula)
                c.number_format = NumFormats.USD_MILLIONS
                c.alignment = Alignment(horizontal="right")
                is_total = row_num == 8
                c.font = Font(name="Calibri", size=11, bold=is_total)
                c.fill = Fills.subheader() if is_total else Fills.white()
                c.border = Borders.thick_bottom() if is_total else Borders.bottom_only()

        # Average balance for interest
        fmt.apply_header_row(10, 2, dc + n - 1, "INTEREST")
        fmt.apply_label_cell(11, 2, "Average Debt Balance", indent=1)
        fmt.apply_label_cell(12, 2, "Interest Rate", indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            avg = f"=({cl}8+{cl}6)/2" if i == 0 else f"=({cl}8+{prev_cl}8)/2"
            ws.cell(row=11, column=col, value=avg).number_format = NumFormats.USD_MILLIONS
            ws.cell(row=12, column=col, value=f"={asm}!C11").number_format = NumFormats.PERCENT_ONE

            for r in [11, 12]:
                c = ws.cell(row=r, column=col)
                c.font = Fonts.body()
                c.fill = Fills.white()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

        # Row 5 = opening balance for IS interest formula
        fmt.apply_label_cell(5, 2, "Average Balance (for Interest)", indent=0)
        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            c = ws.cell(row=5, column=col, value=f"={cl}11")
            c.number_format = NumFormats.USD_MILLIONS
            c.font = Fonts.body()
            c.border = Borders.bottom_only()
            c.alignment = Alignment(horizontal="right")

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_cash_flow_statement(self):
        ws, fmt = self.add_sheet("CASH_FLOW_STMT", tab_color="375623")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"
        is_s = "INCOME_STATEMENT"
        wc = "WORKING_CAPITAL"
        ds = "DEBT_SCHEDULE"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "CASH FLOW STATEMENT", "Indirect Method — Fully Linked")
        fmt.apply_units_label(3, 2, "$ in Millions")
        self._yr_header(ws, 4, n, dc)

        sections = [
            (5, "OPERATING ACTIVITIES", [
                (6,  "Net Income",              f"={is_s}!{{cl}}21",            False),
                (7,  "Add: D&A",                f"=-{is_s}!{{cl}}14",           False),
                (8,  "Change in Working Capital",f"=-{wc}!{{cl}}16",            False),
                (9,  "Cash from Operations",    f"={{cl}}6+{{cl}}7+{{cl}}8",    True),
            ]),
            (11, "INVESTING ACTIVITIES", [
                (12, "Capital Expenditures",    f"=-{is_s}!{{cl}}6*{asm}!C15", False),
                (13, "Cash from Investing",     f"={{cl}}12",                   True),
            ]),
            (15, "FINANCING ACTIVITIES", [
                (16, "Debt Repayment",          f"={ds}!{{cl}}7",               False),
                (17, "Dividends Paid",          f"={is_s}!{{cl}}24",            False),
                (18, "Cash from Financing",     f"={{cl}}16+{{cl}}17",          True),
            ]),
            (20, "NET CASH MOVEMENT", [
                (21, "Net Increase / (Decrease) in Cash", f"={{cl}}9+{{cl}}13+{{cl}}18", True),
            ]),
        ]

        for header_row, section_label, items in sections:
            fmt.apply_header_row(header_row, 2, dc + n - 1, section_label)
            for row_num, label, formula_tpl, is_total in items:
                fmt.apply_label_cell(row_num, 2, label, indent=0 if is_total else 1)

                for i in range(n):
                    col = dc + i
                    cl = get_column_letter(col)
                    formula = formula_tpl.replace("{cl}", cl)
                    c = ws.cell(row=row_num, column=col, value=formula)
                    c.number_format = NumFormats.USD_MILLIONS
                    c.alignment = Alignment(horizontal="right")
                    if is_total:
                        c.font = Font(name="Calibri", size=11, bold=True)
                        c.fill = Fills.subheader()
                        c.border = Borders.thick_bottom()
                    else:
                        c.font = Fonts.body()
                        c.fill = Fills.white()
                        c.border = Borders.bottom_only()

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_balance_sheet(self):
        ws, fmt = self.add_sheet("BALANCE_SHEET", tab_color="2E75B6")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"
        is_s = "INCOME_STATEMENT"
        wc = "WORKING_CAPITAL"
        ds = "DEBT_SCHEDULE"
        cfs = "CASH_FLOW_STMT"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "BALANCE SHEET", "Fully Linked — Checks on CHECKS Sheet")
        fmt.apply_units_label(3, 2, "$ in Millions")
        self._yr_header(ws, 4, n, dc)

        bs_sections = [
            (5, "ASSETS", [
                (6,  "Cash & Equivalents",     True),
                (7,  "Accounts Receivable",    False),
                (8,  "Inventory",              False),
                (9,  "Total Current Assets",   True),
                (11, "PP&E (Net)",             False),
                (12, "Total Assets",           True),
            ]),
            (14, "LIABILITIES", [
                (15, "Accounts Payable",       False),
                (16, "Total Current Liabilities", True),
                (18, "Long-Term Debt",         False),
                (19, "Total Liabilities",      True),
            ]),
            (21, "EQUITY", [
                (22, "Common Equity",          False),
                (23, "Retained Earnings",      False),
                (24, "Total Equity",           True),
                (25, "Total Liabilities + Equity", True),
            ]),
        ]

        for header_row, section_label, items in bs_sections:
            fmt.apply_header_row(header_row, 2, dc + n - 1, section_label)
            for row_num, label, is_bold in items:
                fmt.apply_label_cell(row_num, 2, label, indent=0 if is_bold else 1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                cash   = f"={asm}!C21+{cfs}!{cl}21"
                re     = f"={is_s}!{cl}25"
                equity = f"={asm}!C24"
            else:
                cash   = f"={prev_cl}6+{cfs}!{cl}21"
                re     = f"={prev_cl}23+{is_s}!{cl}25"
                equity = f"={prev_cl}22"

            ppe_formula = f"={asm}!C23+INCOME_STATEMENT!{cl}6*{asm}!C15-(-INCOME_STATEMENT!{cl}14)" if i == 0 else \
                          f"={prev_cl}11+INCOME_STATEMENT!{cl}6*{asm}!C15-(-INCOME_STATEMENT!{cl}14)"

            formulas = {
                6:  cash,
                7:  f"={wc}!{cl}10",
                8:  f"={wc}!{cl}11",
                9:  f"={cl}6+{cl}7+{cl}8",
                11: ppe_formula,
                12: f"={cl}9+{cl}11",
                15: f"={wc}!{cl}12",
                16: f"={cl}15",
                18: f"={ds}!{cl}8",
                19: f"={cl}16+{cl}18",
                22: equity,
                23: re,
                24: f"={cl}22+{cl}23",
                25: f"={cl}19+{cl}24",
            }
            total_rows = [9, 12, 16, 19, 24, 25]

            for row_num, formula in formulas.items():
                c = ws.cell(row=row_num, column=col, value=formula)
                c.number_format = NumFormats.USD_MILLIONS
                c.alignment = Alignment(horizontal="right")
                if row_num in total_rows:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_checks(self):
        ws, fmt = self.add_sheet("CHECKS", tab_color="C00000")
        n = self.years
        dc = self.dc

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "MODEL INTEGRITY CHECKS", "All checks must show ✓ — Red = ERROR")
        self._yr_header(ws, 4, n, dc)

        fmt.apply_header_row(5, 2, dc + n - 1, "BALANCE SHEET CHECK")
        fmt.apply_label_cell(6, 2, "Total Assets")
        fmt.apply_label_cell(7, 2, "Total Liabilities + Equity")
        fmt.apply_label_cell(8, 2, "Difference (must = 0)")
        fmt.apply_label_cell(9, 2, "BS Check")

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)

            ws.cell(row=6, column=col, value=f"=BALANCE_SHEET!{cl}12").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=7, column=col, value=f"=BALANCE_SHEET!{cl}25").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=8, column=col, value=f"=ROUND({cl}6-{cl}7,2)").number_format = NumFormats.USD_MILLIONS

            check = ws.cell(row=9, column=col,
                value=f"=IF(ABS({cl}8)<0.01,\"✓ BALANCED\",\"✗ OUT OF BALANCE\")")
            check.font = Font(name="Calibri", size=10, bold=True)
            check.alignment = Alignment(horizontal="center")

            for r in [6, 7, 8]:
                c = ws.cell(row=r, column=col)
                c.font = Fonts.body()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

        fmt.freeze_panes("A1")

    def _build_dashboard(self):
        ws, fmt = self.add_sheet("DASHBOARD", tab_color=Colors.ACCENT_GOLD)
        n = self.years
        dc = self.dc
        last_cl = get_column_letter(dc + n - 1)
        asm = "ASSUMPTIONS"
        is_s = "INCOME_STATEMENT"
        bs = "BALANCE_SHEET"

        ws.sheet_view.showGridLines = False
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["A"].width = 2

        for row in range(1, 4):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)

        ws.merge_cells("B2:O2")
        t = ws["B2"]
        t.value = f"3-STATEMENT DASHBOARD — {self.a.get('company_name','Company')}"
        t.font = Font(name="Calibri", size=16, bold=True, color=Colors.WHITE)
        t.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 30

        # KPI cards
        kpis = [
            ("B", "Net Revenue ($M)",  f"={is_s}!{last_cl}6",   NumFormats.USD_MILLIONS),
            ("E", "EBITDA ($M)",       f"={is_s}!{last_cl}13",  NumFormats.USD_MILLIONS),
            ("H", "Net Income ($M)",   f"={is_s}!{last_cl}21",  NumFormats.USD_MILLIONS),
            ("K", "Total Assets ($M)", f"={bs}!{last_cl}12",    NumFormats.USD_MILLIONS),
            ("N", "Net Debt ($M)",     f"={bs}!{last_cl}18-{bs}!{last_cl}6", NumFormats.USD_MILLIONS),
        ]

        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 38

        for col_letter, label, formula, fmt_code in kpis:
            lbl = ws[f"{col_letter}5"]
            lbl.value = label
            lbl.font = Fonts.kpi_label()
            lbl.fill = Fills.kpi_card()
            lbl.alignment = Alignment(horizontal="center", vertical="center")
            lbl.border = Border(
                top=Side(style="medium", color=Colors.INST_BLUE),
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE)
            )
            val = ws[f"{col_letter}6"]
            val.value = formula
            val.font = Fonts.kpi_value()
            val.number_format = fmt_code
            val.fill = Fills.kpi_card()
            val.alignment = Alignment(horizontal="center", vertical="center")
            val.border = Border(
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE),
                bottom=Side(style="medium", color=Colors.INST_BLUE)
            )

        # Chart data
        cr = 50
        for i in range(n):
            ws.cell(row=cr, column=3 + i, value=f"FY{self.base_year+i+1}")
            ws.cell(row=cr + 1, column=3 + i, value=f"={is_s}!{get_column_letter(dc+i)}6")
            ws.cell(row=cr + 2, column=3 + i, value=f"={is_s}!{get_column_letter(dc+i)}13")
            ws.cell(row=cr + 3, column=3 + i, value=f"={is_s}!{get_column_letter(dc+i)}21")

        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue / EBITDA / Net Income ($M)"
        chart.style = 10
        chart.width = 28
        chart.height = 12

        for ref_row, name in [(cr+1, "Revenue"), (cr+2, "EBITDA"), (cr+3, "Net Income")]:
            ref = Reference(ws, min_col=3, max_col=2+n, min_row=ref_row)
            chart.add_data(ref)

        chart.set_categories(Reference(ws, min_col=3, max_col=2+n, min_row=cr))
        for i, name in enumerate(["Revenue", "EBITDA", "Net Income"]):
            chart.series[i].title = SeriesLabel(v=name)

        ws.add_chart(chart, "B8")
        fmt.hide_gridlines()
