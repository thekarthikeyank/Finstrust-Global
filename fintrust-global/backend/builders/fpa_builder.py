"""
FP&A Forecast Model Builder — 8 Sheets
Actual vs Budget vs Forecast with Variance Analysis
"""

from builders.base_builder import BaseBuilder
from formatting.institutional import Colors, Fonts, Fills, Borders, NumFormats, ColWidths, Formatter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


class FPABuilder(BaseBuilder):

    def __init__(self, assumptions: dict):
        super().__init__(assumptions)
        self.a = assumptions
        self.years = int(self.a.get("projection_years", 3))
        self.base_year = int(self.a.get("base_year", 2024))
        self.dc = 3
        self.months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def build(self):
        self.build_cover("FP&A Forecast")
        self._build_assumptions()
        self._build_revenue_build()
        self._build_opex_build()
        self._build_pl_summary()
        self._build_variance_analysis()
        self._build_rolling_forecast()
        self._build_dashboard()
        return self.save_to_buffer()

    def _build_assumptions(self):
        ws, fmt = self.add_sheet("ASSUMPTIONS", tab_color="2E75B6")
        a = self.a

        fmt.set_column_widths({"A": 3, "B": 40, "C": 22, "D": 22})
        fmt.apply_sheet_title(2, 2, "FP&A ASSUMPTIONS", "Budget & Forecast Drivers")
        fmt.apply_units_label(3, 2, "$ in Millions")

        sections = [
            (5, "REVENUE ASSUMPTIONS", [
                ("Total Revenue Budget ($M)",   "total_revenue_budget", 120.0, NumFormats.USD_MILLIONS),
                ("Revenue Growth vs PY %",      "rev_growth",           0.10,  NumFormats.PERCENT_ONE),
                ("Q1 Revenue Weight %",         "q1_weight",            0.22,  NumFormats.PERCENT_ONE),
                ("Q2 Revenue Weight %",         "q2_weight",            0.25,  NumFormats.PERCENT_ONE),
                ("Q3 Revenue Weight %",         "q3_weight",            0.24,  NumFormats.PERCENT_ONE),
                ("Q4 Revenue Weight %",         "q4_weight",            0.29,  NumFormats.PERCENT_ONE),
            ]),
            (13, "COST ASSUMPTIONS", [
                ("Gross Margin % (Budget)",     "gross_margin",         0.58,  NumFormats.PERCENT_ONE),
                ("S&M as % Revenue",            "sm_pct",               0.15,  NumFormats.PERCENT_ONE),
                ("R&D as % Revenue",            "rd_pct",               0.12,  NumFormats.PERCENT_ONE),
                ("G&A as % Revenue",            "ga_pct",               0.08,  NumFormats.PERCENT_ONE),
                ("D&A ($M, fixed)",             "da_amount",            5.0,   NumFormats.USD_MILLIONS),
            ]),
            (20, "HEADCOUNT", [
                ("Total Headcount (Budget)",    "hc_budget",            150,   NumFormats.INTEGER),
                ("Avg Fully Loaded Cost ($k)",  "hc_cost_k",            120.0, NumFormats.USD_MILLIONS),
                ("Headcount Growth %",          "hc_growth",            0.08,  NumFormats.PERCENT_ONE),
            ]),
            (25, "ACTUALS (YTD)", [
                ("Actual Revenue YTD ($M)",     "actual_rev_ytd",       60.0,  NumFormats.USD_MILLIONS),
                ("Budget Revenue YTD ($M)",     "budget_rev_ytd",       58.0,  NumFormats.USD_MILLIONS),
                ("Actual EBITDA YTD ($M)",      "actual_ebitda_ytd",    13.0,  NumFormats.USD_MILLIONS),
                ("Budget EBITDA YTD ($M)",      "budget_ebitda_ytd",    12.5,  NumFormats.USD_MILLIONS),
            ]),
        ]

        for header_row, section_label, items in sections:
            fmt.apply_header_row(header_row, 2, 4, section_label)
            for i, (lbl, key, default, fmt_code) in enumerate(items):
                fmt.apply_label_cell(header_row + 1 + i, 2, lbl)
                val = float(a.get(key, default))
                fmt.apply_input_cell(header_row + 1 + i, 3, val, fmt_code)

    def _build_revenue_build(self):
        ws, fmt = self.add_sheet("REVENUE_BUILD", tab_color="375623")
        asm = "ASSUMPTIONS"

        # Monthly columns
        fmt.set_column_widths({"A": 3, "B": 32})
        for i in range(12):
            ws.column_dimensions[get_column_letter(self.dc + i)].width = 10
        ws.column_dimensions[get_column_letter(self.dc + 12)].width = 14  # Total col

        fmt.apply_sheet_title(2, 2, "REVENUE BUILD", f"Monthly Budget vs Forecast — FY{self.base_year + 1}")
        fmt.apply_units_label(3, 2, "$ in Millions")

        # Month headers
        for i, month in enumerate(self.months):
            col = self.dc + i
            cell = ws.cell(row=4, column=col, value=month)
            cell.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        total_col = self.dc + 12
        ws.cell(row=4, column=total_col, value="Full Year").font = Font(
            name="Calibri", size=10, bold=True, color=Colors.WHITE)
        ws.cell(row=4, column=total_col).fill = PatternFill("solid", fgColor=Colors.INST_BLUE)
        ws.cell(row=4, column=total_col).alignment = Alignment(horizontal="center")

        # Quarter weights for monthly split
        q_weights = {
            "q1": [1/3, 1/3, 1/3],  # Jan Feb Mar
            "q2": [1/3, 1/3, 1/3],
            "q3": [1/3, 1/3, 1/3],
            "q4": [1/3, 1/3, 1/3],
        }

        # Revenue rows
        fmt.apply_header_row(5, 2, total_col, "REVENUE BUDGET")
        segments = [
            ("Segment A (Product)",   0.50),
            ("Segment B (Services)",  0.30),
            ("Segment C (Other)",     0.20),
        ]
        quarter_keys = [f"{asm}!C8", f"{asm}!C9", f"{asm}!C10", f"{asm}!C11"]

        for seg_i, (seg_name, seg_weight) in enumerate(segments):
            row = 6 + seg_i
            fmt.apply_label_cell(row, 2, seg_name, indent=1)
            row_vals = []
            for m_i in range(12):
                col = self.dc + m_i
                q_idx = m_i // 3
                monthly_formula = f"={asm}!C6*{seg_weight}*{quarter_keys[q_idx]}/3"
                c = ws.cell(row=row, column=col, value=monthly_formula)
                c.number_format = NumFormats.USD_MILLIONS
                c.font = Fonts.body()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")
                row_vals.append(get_column_letter(col) + str(row))

            # Total
            tc = ws.cell(row=row, column=total_col,
                value=f"=SUM({get_column_letter(self.dc)}{row}:{get_column_letter(self.dc+11)}{row})")
            tc.number_format = NumFormats.USD_MILLIONS
            tc.font = Font(name="Calibri", size=10, bold=True)
            tc.fill = Fills.subheader()
            tc.border = Borders.thick_bottom()
            tc.alignment = Alignment(horizontal="right")

        # Total Revenue row
        total_row = 9
        fmt.apply_total_row(total_row, 2, total_col, "Total Revenue Budget", NumFormats.USD_MILLIONS)
        for m_i in range(12):
            col = self.dc + m_i
            cl = get_column_letter(col)
            c = ws.cell(row=total_row, column=col, value=f"=SUM({cl}6:{cl}8)")
            c.number_format = NumFormats.USD_MILLIONS
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = Fills.subheader()
            c.border = Borders.thick_bottom()
            c.alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=total_col,
            value=f"=SUM({get_column_letter(self.dc)}{total_row}:{get_column_letter(self.dc+11)}{total_row})"
        ).number_format = NumFormats.USD_MILLIONS

        fmt.freeze_panes(f"{get_column_letter(self.dc)}5")

    def _build_opex_build(self):
        ws, fmt = self.add_sheet("OPEX_BUILD", tab_color="7030A0")
        asm = "ASSUMPTIONS"
        rb = "REVENUE_BUILD"

        fmt.set_column_widths({"A": 3, "B": 36})
        for i in range(12):
            ws.column_dimensions[get_column_letter(self.dc + i)].width = 10
        ws.column_dimensions[get_column_letter(self.dc + 12)].width = 14

        fmt.apply_sheet_title(2, 2, "OPEX BUILD", f"Operating Expense Budget — FY{self.base_year + 1}")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i, month in enumerate(self.months):
            col = self.dc + i
            cell = ws.cell(row=4, column=col, value=month)
            cell.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=4, column=self.dc + 12, value="Full Year").font = Font(
            name="Calibri", size=10, bold=True, color=Colors.WHITE)
        ws.cell(row=4, column=self.dc + 12).fill = PatternFill("solid", fgColor=Colors.INST_BLUE)
        ws.cell(row=4, column=self.dc + 12).alignment = Alignment(horizontal="center")

        total_col = self.dc + 12
        fmt.apply_header_row(5, 2, total_col, "OPERATING EXPENSES")
        opex_lines = [
            (6,  "Sales & Marketing",    f"={rb}!{{cl}}9*{asm}!C14"),
            (7,  "Research & Development", f"={rb}!{{cl}}9*{asm}!C15"),
            (8,  "General & Administrative", f"={rb}!{{cl}}9*{asm}!C16"),
            (9,  "D&A",                  f"={asm}!C18/12"),
            (10, "Headcount Cost",       f"={asm}!C21*{asm}!C22/1000/12"),
        ]

        for row_num, label, formula_tpl in opex_lines:
            fmt.apply_label_cell(row_num, 2, label, indent=1)
            for m_i in range(12):
                col = self.dc + m_i
                cl = get_column_letter(col)
                formula = formula_tpl.replace("{cl}", cl)
                c = ws.cell(row=row_num, column=col, value=formula)
                c.number_format = NumFormats.USD_MILLIONS
                c.font = Fonts.body()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

            tc = ws.cell(row=row_num, column=total_col,
                value=f"=SUM({get_column_letter(self.dc)}{row_num}:{get_column_letter(self.dc+11)}{row_num})")
            tc.number_format = NumFormats.USD_MILLIONS
            tc.font = Font(name="Calibri", size=10, bold=True)
            tc.fill = Fills.subheader()
            tc.border = Borders.thick_bottom()
            tc.alignment = Alignment(horizontal="right")

        # Total OpEx
        fmt.apply_total_row(11, 2, total_col, "Total Operating Expenses", NumFormats.USD_MILLIONS)
        for m_i in range(12):
            col = self.dc + m_i
            cl = get_column_letter(col)
            c = ws.cell(row=11, column=col, value=f"=SUM({cl}6:{cl}10)")
            c.number_format = NumFormats.USD_MILLIONS
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = Fills.subheader()
            c.border = Borders.thick_bottom()
            c.alignment = Alignment(horizontal="right")

        fmt.freeze_panes(f"{get_column_letter(self.dc)}5")

    def _build_pl_summary(self):
        ws, fmt = self.add_sheet("PL_SUMMARY", tab_color="375623")
        asm = "ASSUMPTIONS"
        rb = "REVENUE_BUILD"
        ob = "OPEX_BUILD"

        # Annual columns: Actual, Budget, Forecast for each year
        fmt.set_column_widths({"A": 3, "B": 36, "C": 18, "D": 18, "E": 18, "F": 5})
        fmt.apply_sheet_title(2, 2, "P&L SUMMARY", "Actual vs Budget vs Forecast")
        fmt.apply_units_label(3, 2, "$ in Millions")

        # Column headers
        headers = [
            (3, "ACTUAL", Colors.POSITIVE_GREEN),
            (4, "BUDGET", Colors.INST_BLUE),
            (5, "FORECAST", Colors.ACCENT_GOLD),
        ]
        for col, label, color in headers:
            c = ws.cell(row=4, column=col, value=label)
            c.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center")

        fmt.apply_header_row(5, 2, 5, "INCOME STATEMENT SUMMARY")
        pl_lines = [
            (6,  "Total Revenue",    f"={asm}!C25",  f"={rb}!{get_column_letter(self.dc+12)}9",  f"={rb}!{get_column_letter(self.dc+12)}9*1.02"),
            (7,  "Gross Profit",     f"={asm}!C25*{asm}!C13",  f"={rb}!{get_column_letter(self.dc+12)}9*{asm}!C13", f"={rb}!{get_column_letter(self.dc+12)}9*1.02*{asm}!C13"),
            (8,  "Total OpEx",       f"=-{asm}!C26",  f"={ob}!{get_column_letter(self.dc+12)}11", f"={ob}!{get_column_letter(self.dc+12)}11*1.01"),
            (9,  "EBITDA",           f"={asm}!C27",  f"=C7+C8",  f"=E7+E8"),
            (10, "D&A",              f"=-{asm}!C18",  f"=-{asm}!C18",  f"=-{asm}!C18"),
            (11, "EBIT",             f"=C9+C10",  f"=D9+D10",  f"=E9+E10"),
        ]

        for row_num, label, actual, budget, forecast in pl_lines:
            fmt.apply_label_cell(row_num, 2, label, indent=1)
            for col, val in [(3, actual), (4, budget), (5, forecast)]:
                c = ws.cell(row=row_num, column=col, value=val)
                c.number_format = NumFormats.USD_MILLIONS
                c.alignment = Alignment(horizontal="right")
                if row_num in [9, 11]:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()

        # Margin summary
        fmt.apply_header_row(13, 2, 5, "MARGIN ANALYSIS")
        for row_num, label, row_ref in [
            (14, "Gross Margin %", 7), (15, "EBITDA Margin %", 9), (16, "EBIT Margin %", 11)
        ]:
            fmt.apply_label_cell(row_num, 2, label, indent=1)
            for col in [3, 4, 5]:
                cl = get_column_letter(col)
                c = ws.cell(row=row_num, column=col, value=f"={cl}{row_ref}/{cl}6")
                c.number_format = NumFormats.PERCENT_ONE
                c.font = Fonts.body()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

    def _build_variance_analysis(self):
        ws, fmt = self.add_sheet("VARIANCE", tab_color="C00000")
        pl = "PL_SUMMARY"

        fmt.set_column_widths({
            "A": 3, "B": 32, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14
        })
        fmt.apply_sheet_title(2, 2, "VARIANCE ANALYSIS", "Actual vs Budget — $ and % Variance + RAG Status")
        fmt.apply_units_label(3, 2, "$ in Millions")

        headers = [
            (3, "ACTUAL"),  (4, "BUDGET"),
            (5, "VAR ($)"), (6, "VAR (%)"), (7, "STATUS")
        ]
        for col, label in headers:
            c = ws.cell(row=4, column=col, value=label)
            c.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            c.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            c.alignment = Alignment(horizontal="center")

        fmt.apply_header_row(5, 2, 7, "P&L VARIANCE — YTD")
        lines = [
            (6,  "Revenue",      "C6", "D6",  True),
            (7,  "Gross Profit", "C7", "D7",  True),
            (8,  "Total OpEx",   "C8", "D8",  False),
            (9,  "EBITDA",       "C9", "D9",  True),
            (11, "EBIT",         "C11","D11", True),
        ]

        for row_num, label, actual_ref, budget_ref, favorable_positive in lines:
            fmt.apply_label_cell(row_num, 2, label, indent=1)
            ws.cell(row=row_num, column=3, value=f"={pl}!{actual_ref}").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=row_num, column=4, value=f"={pl}!{budget_ref}").number_format = NumFormats.USD_MILLIONS

            var_usd = ws.cell(row=row_num, column=5,
                value=f"={pl}!{actual_ref}-{pl}!{budget_ref}")
            var_usd.number_format = NumFormats.USD_MILLIONS
            var_usd.alignment = Alignment(horizontal="right")

            var_pct = ws.cell(row=row_num, column=6,
                value=f"=IFERROR(({pl}!{actual_ref}-{pl}!{budget_ref})/{pl}!{budget_ref},0)")
            var_pct.number_format = NumFormats.PERCENT_ONE
            var_pct.alignment = Alignment(horizontal="right")

            # RAG status
            if favorable_positive:
                status_formula = f'=IF(E{row_num}>0,"🟢 ABOVE","🔴 BELOW")'
            else:
                status_formula = f'=IF(E{row_num}<0,"🟢 BELOW","🔴 ABOVE")'

            status = ws.cell(row=row_num, column=7, value=status_formula)
            status.font = Font(name="Calibri", size=10)
            status.alignment = Alignment(horizontal="center")

            for c_col in [3, 4, 5, 6]:
                c = ws.cell(row=row_num, column=c_col)
                c.font = Fonts.body()
                c.border = Borders.bottom_only()
                c.fill = Fills.white()

    def _build_rolling_forecast(self):
        ws, fmt = self.add_sheet("ROLLING_FORECAST", tab_color="7030A0")
        asm = "ASSUMPTIONS"
        rb = "REVENUE_BUILD"

        fmt.set_column_widths({"A": 3, "B": 32})
        for i in range(12):
            ws.column_dimensions[get_column_letter(self.dc + i)].width = 10
        ws.column_dimensions[get_column_letter(self.dc + 12)].width = 14

        fmt.apply_sheet_title(2, 2, "ROLLING 12-MONTH FORECAST", "Budget vs Forecast by Month")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i, month in enumerate(self.months):
            col = self.dc + i
            c = ws.cell(row=4, column=col, value=month)
            c.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            c.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            c.alignment = Alignment(horizontal="center")

        fmt.apply_header_row(5, 2, self.dc + 12, "REVENUE")
        rolling_rows = [
            (6,  "Budget Revenue",   f"={rb}!{{cl}}9",        False),
            (7,  "Forecast Revenue", f"={rb}!{{cl}}9*1.02",   False),
            (8,  "Actual (if available)", f"=IF({{m_i}}<6,{rb}!{{cl}}9*RAND()*0.1+{rb}!{{cl}}9,\"\")", False),
            (9,  "Variance ($)",     f"=IFERROR({{cl}}7-{{cl}}6,0)", True),
        ]

        for row_num, label, formula_tpl, is_total in rolling_rows:
            fmt.apply_label_cell(row_num, 2, label, indent=0 if is_total else 1)
            for m_i in range(12):
                col = self.dc + m_i
                cl = get_column_letter(col)
                formula = formula_tpl.replace("{cl}", cl).replace("{m_i}", str(m_i))
                c = ws.cell(row=row_num, column=col, value=formula)
                c.number_format = NumFormats.USD_MILLIONS
                c.alignment = Alignment(horizontal="right")
                if is_total:
                    c.font = Font(name="Calibri", size=10, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()

        fmt.freeze_panes(f"{get_column_letter(self.dc)}5")

    def _build_dashboard(self):
        ws, fmt = self.add_sheet("DASHBOARD", tab_color=Colors.ACCENT_GOLD)
        asm = "ASSUMPTIONS"
        pl = "PL_SUMMARY"
        var = "VARIANCE"

        ws.sheet_view.showGridLines = False
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["A"].width = 2

        for row in range(1, 4):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)

        ws.merge_cells("B2:O2")
        t = ws["B2"]
        t.value = f"FP&A DASHBOARD — {self.a.get('company_name','Company')} | FY{self.base_year + 1}"
        t.font = Font(name="Calibri", size=16, bold=True, color=Colors.WHITE)
        t.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 30

        kpis = [
            ("B", "Actual Rev ($M)",    f"={pl}!C6",  NumFormats.USD_MILLIONS),
            ("E", "Budget Rev ($M)",    f"={pl}!D6",  NumFormats.USD_MILLIONS),
            ("H", "Rev Variance ($M)",  f"={var}!E6", NumFormats.USD_MILLIONS),
            ("K", "Actual EBITDA ($M)", f"={pl}!C9",  NumFormats.USD_MILLIONS),
            ("N", "EBITDA Margin",      f"=IFERROR({pl}!C9/{pl}!C6,0)", NumFormats.PERCENT_ONE),
        ]

        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 38

        for col_letter, label, formula, fmt_code in kpis:
            lbl = ws[f"{col_letter}5"]
            lbl.value = label
            lbl.font = Fonts.kpi_label()
            lbl.fill = Fills.kpi_card()
            lbl.alignment = Alignment(horizontal="center", vertical="center")
            lbl.border = Border(
                top=Side(style="medium", color=Colors.INST_BLUE),
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE)
            )
            val = ws[f"{col_letter}6"]
            val.value = formula
            val.font = Fonts.kpi_value()
            val.number_format = fmt_code
            val.fill = Fills.kpi_card()
            val.alignment = Alignment(horizontal="center", vertical="center")
            val.border = Border(
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE),
                bottom=Side(style="medium", color=Colors.INST_BLUE)
            )

        # Chart
        cr = 50
        for i, month in enumerate(self.months):
            ws.cell(row=cr, column=3 + i, value=month)
            ws.cell(row=cr + 1, column=3 + i, value=f"=ROLLING_FORECAST!{get_column_letter(self.dc+i)}6")
            ws.cell(row=cr + 2, column=3 + i, value=f"=ROLLING_FORECAST!{get_column_letter(self.dc+i)}7")

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Budget vs Forecast Revenue ($M)"
        chart.style = 10
        chart.width = 28
        chart.height = 12

        for ref_row, name in [(cr+1, "Budget"), (cr+2, "Forecast")]:
            ref = Reference(ws, min_col=3, max_col=14, min_row=ref_row)
            chart.add_data(ref)
        chart.set_categories(Reference(ws, min_col=3, max_col=14, min_row=cr))
        from openpyxl.chart.series import SeriesLabel
        chart.series[0].title = SeriesLabel(v="Budget")
        chart.series[1].title = SeriesLabel(v="Forecast")
        ws.add_chart(chart, "B8")

        fmt.hide_gridlines()
