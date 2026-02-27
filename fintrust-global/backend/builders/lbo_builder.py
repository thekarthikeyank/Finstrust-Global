"""
LBO Model Builder — 8 Sheet Institutional Model
Sheets: COVER, ASSUMPTIONS, SOURCES_USES, INCOME_STATEMENT,
        DEBT_SCHEDULE, CASH_FLOW, RETURNS, DASHBOARD
"""

from builders.base_builder import BaseBuilder
from formatting.institutional import Colors, Fonts, Fills, Borders, NumFormats, ColWidths, Formatter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


class LBOBuilder(BaseBuilder):

    def __init__(self, assumptions: dict):
        super().__init__(assumptions)
        self.a = assumptions
        self.years = int(self.a.get("hold_period", 5))
        self.base_year = int(self.a.get("base_year", 2024))
        self.dc = 3

    def build(self):
        self.build_cover("LBO")
        self._build_assumptions()
        self._build_sources_uses()
        self._build_income_statement()
        self._build_debt_schedule()
        self._build_cash_flow()
        self._build_returns()
        self._build_dashboard()
        return self.save_to_buffer()

    def _build_assumptions(self):
        ws, fmt = self.add_sheet("ASSUMPTIONS", tab_color="2E75B6")
        a = self.a

        fmt.set_column_widths({"A": 3, "B": 40, "C": 22, "D": 22})
        fmt.apply_sheet_title(2, 2, "LBO ASSUMPTIONS", "All blue cells are editable inputs")
        fmt.apply_units_label(3, 2, "$ in Millions unless stated")

        fmt.apply_header_row(5, 2, 4, "TRANSACTION ASSUMPTIONS")
        txn_rows = [
            ("Entry EBITDA ($M)",      "entry_ebitda",     50.0,  NumFormats.USD_MILLIONS),
            ("Entry EV/EBITDA Multiple","entry_multiple",   8.0,   NumFormats.MULTIPLE),
            ("Entry Enterprise Value", "entry_ev",         400.0, NumFormats.USD_MILLIONS),
            ("Transaction Fees %",    "txn_fees_pct",      0.02,  NumFormats.PERCENT_ONE),
            ("Hold Period (Years)",   "hold_period",       5,     NumFormats.INTEGER),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(txn_rows):
            fmt.apply_label_cell(6 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(6 + i, 3, val, fmt_code)

        fmt.apply_header_row(12, 2, 4, "DEBT STRUCTURE")
        debt_rows = [
            ("Senior Debt / EBITDA",    "senior_leverage",  4.0,  NumFormats.MULTIPLE),
            ("Senior Interest Rate",    "senior_rate",      0.07, NumFormats.PERCENT_ONE),
            ("Subordinated Debt / EBITDA","sub_leverage",   1.5,  NumFormats.MULTIPLE),
            ("Sub Debt Interest Rate",  "sub_rate",         0.10, NumFormats.PERCENT_ONE),
            ("Debt Amortization % p.a.","amort_pct",        0.05, NumFormats.PERCENT_ONE),
            ("Cash Sweep %",            "cash_sweep_pct",   0.50, NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(debt_rows):
            fmt.apply_label_cell(13 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(13 + i, 3, val, fmt_code)

        fmt.apply_header_row(20, 2, 4, "OPERATING ASSUMPTIONS")
        ops_rows = [
            ("Revenue Year 1 ($M)",      "base_revenue",    200.0, NumFormats.USD_MILLIONS),
            ("Revenue Growth Rate",      "rev_growth",      0.08,  NumFormats.PERCENT_ONE),
            ("EBITDA Margin %",          "ebitda_margin",   0.25,  NumFormats.PERCENT_ONE),
            ("D&A as % Revenue",         "da_pct",          0.05,  NumFormats.PERCENT_ONE),
            ("Capex as % Revenue",       "capex_pct",       0.06,  NumFormats.PERCENT_ONE),
            ("Tax Rate",                 "tax_rate",        0.25,  NumFormats.PERCENT_ONE),
            ("Change in NWC as % Rev",   "nwc_pct",         0.02,  NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(ops_rows):
            fmt.apply_label_cell(21 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(21 + i, 3, val, fmt_code)

        fmt.apply_header_row(29, 2, 4, "EXIT ASSUMPTIONS")
        exit_rows = [
            ("Exit EV/EBITDA Multiple",  "exit_multiple",   9.0,  NumFormats.MULTIPLE),
            ("Management Rollover %",    "mgmt_rollover",   0.0,  NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(exit_rows):
            fmt.apply_label_cell(30 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(30 + i, 3, val, fmt_code)

    def _build_sources_uses(self):
        ws, fmt = self.add_sheet("SOURCES_USES", tab_color="7030A0")
        asm = "ASSUMPTIONS"

        fmt.set_column_widths({"A": 3, "B": 36, "C": 20, "D": 14, "E": 3, "F": 36, "G": 20, "H": 14})
        fmt.apply_sheet_title(2, 2, "SOURCES & USES", "Transaction Financing Structure")
        fmt.apply_units_label(3, 2, "$ in Millions")

        # USES (left side)
        fmt.apply_header_row(5, 2, 4, "USES OF FUNDS")
        uses = [
            (6,  "Purchase Price (EV)",    f"={asm}!C7",                       NumFormats.USD_MILLIONS),
            (7,  "Transaction Fees",       f"={asm}!C7*{asm}!C9",              NumFormats.USD_MILLIONS),
            (8,  "Financing Fees",         f"={asm}!C7*0.01",                  NumFormats.USD_MILLIONS),
            (9,  "Total Uses",             f"=C6+C7+C8",                       NumFormats.USD_MILLIONS),
        ]
        for row_num, label, formula, fmt_code in uses:
            fmt.apply_label_cell(row_num, 2, label, indent=1)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.font = Font(name="Calibri", size=11, bold=(row_num == 9))
            c.fill = Fills.subheader() if row_num == 9 else Fills.white()
            c.border = Borders.thick_bottom() if row_num == 9 else Borders.bottom_only()
            c.number_format = fmt_code
            c.alignment = Alignment(horizontal="right")
            pct = ws.cell(row=row_num, column=4, value=f"=C{row_num}/C9" if row_num != 9 else "=1")
            pct.number_format = NumFormats.PERCENT_ONE
            pct.font = Fonts.body()
            pct.alignment = Alignment(horizontal="right")

        # SOURCES (right side)
        fmt.apply_header_row(5, 6, 8, "SOURCES OF FUNDS")
        sources = [
            (6,  "Senior Debt",       f"={asm}!C6*{asm}!C13",  NumFormats.USD_MILLIONS),
            (7,  "Subordinated Debt", f"={asm}!C6*{asm}!C15",  NumFormats.USD_MILLIONS),
            (8,  "Sponsor Equity",    f"=C9-G6-G7",            NumFormats.USD_MILLIONS),
            (9,  "Total Sources",     f"=G6+G7+G8",            NumFormats.USD_MILLIONS),
        ]
        for row_num, label, formula, fmt_code in sources:
            fmt.apply_label_cell(row_num, 6, label, indent=1)
            c = ws.cell(row=row_num, column=7, value=formula)
            c.font = Font(name="Calibri", size=11, bold=(row_num == 9))
            c.fill = Fills.subheader() if row_num == 9 else Fills.white()
            c.border = Borders.thick_bottom() if row_num == 9 else Borders.bottom_only()
            c.number_format = fmt_code
            c.alignment = Alignment(horizontal="right")
            pct = ws.cell(row=row_num, column=8, value=f"=G{row_num}/G9" if row_num != 9 else "=1")
            pct.number_format = NumFormats.PERCENT_ONE
            pct.font = Fonts.body()
            pct.alignment = Alignment(horizontal="right")

        # Check
        ws.cell(row=11, column=2, value="BALANCE CHECK (Sources = Uses):")
        check = ws.cell(row=11, column=3, value="=IF(G9=C9,\"✓ BALANCED\",\"✗ OUT OF BALANCE\")")
        check.font = Font(name="Calibri", size=11, bold=True, color=Colors.POSITIVE_GREEN)

        # Capital Structure Summary
        fmt.apply_header_row(13, 2, 4, "CAPITAL STRUCTURE SUMMARY")
        cap_rows = [
            (14, "Total Debt",        f"=G6+G7",        NumFormats.USD_MILLIONS),
            (15, "Sponsor Equity",    f"=G8",           NumFormats.USD_MILLIONS),
            (16, "Total Capitalization", f"=C14+C15",   NumFormats.USD_MILLIONS),
            (17, "Debt / Total Cap",  f"=C14/C16",      NumFormats.PERCENT_ONE),
            (18, "Equity / Total Cap","=C15/C16",       NumFormats.PERCENT_ONE),
            (19, "Entry Leverage (x)","=C14/ASSUMPTIONS!C6", NumFormats.MULTIPLE),
        ]
        for row_num, label, formula, fmt_code in cap_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.number_format = fmt_code
            c.font = Fonts.body()
            c.border = Borders.bottom_only()
            c.alignment = Alignment(horizontal="right")

    def _build_income_statement(self):
        ws, fmt = self.add_sheet("INCOME_STATEMENT", tab_color="375623")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "INCOME STATEMENT", f"LBO Operating Model — {n}-Year Projection")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"Year {i+1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        fmt.apply_header_row(5, 2, dc + n - 1, "INCOME STATEMENT")

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                rev = f"={asm}!C21"
            else:
                rev = f"={prev_cl}7*(1+{asm}!C22)"

            row_data = {
                7:  (rev,                                   NumFormats.USD_MILLIONS, False, "Net Revenue"),
                8:  (f"=-{cl}7*(1-{asm}!C23)",             NumFormats.USD_MILLIONS, False, "Cost of Goods Sold"),
                9:  (f"={cl}7*{asm}!C23",                  NumFormats.USD_MILLIONS, True,  "EBITDA"),
                10: (f"=-{cl}7*{asm}!C24",                 NumFormats.USD_MILLIONS, False, "D&A"),
                11: (f"={cl}9+{cl}10",                     NumFormats.USD_MILLIONS, True,  "EBIT"),
                12: (f"=-DEBT_SCHEDULE!{cl}7*{asm}!C14+-DEBT_SCHEDULE!{cl}11*{asm}!C16", NumFormats.USD_MILLIONS, False, "Interest Expense"),
                13: (f"={cl}11+{cl}12",                    NumFormats.USD_MILLIONS, False, "EBT"),
                14: (f"=MAX(-{cl}13*{asm}!C27,0)",         NumFormats.USD_MILLIONS, False, "Tax Provision"),
                15: (f"={cl}13+{cl}14",                    NumFormats.USD_MILLIONS, True,  "Net Income"),
                17: (f"={cl}9/{cl}7",                      NumFormats.PERCENT_ONE,  False, "EBITDA Margin %"),
                18: (f"={cl}15/{cl}7",                     NumFormats.PERCENT_ONE,  False, "Net Margin %"),
            }

            for row_num, (formula, fmt_code, is_total, _) in row_data.items():
                c = ws.cell(row=row_num, column=col, value=formula)
                c.number_format = fmt_code
                c.alignment = Alignment(horizontal="right")
                if is_total:
                    c.font = Font(name="Calibri", size=11, bold=True)
                    c.fill = Fills.subheader()
                    c.border = Borders.thick_bottom()
                else:
                    c.font = Fonts.body()
                    c.fill = Fills.white()
                    c.border = Borders.bottom_only()

        # Labels
        for row_num, label, indent in [
            (7, "Net Revenue", 1), (8, "COGS", 1), (9, "EBITDA", 0),
            (10, "D&A", 1), (11, "EBIT", 0), (12, "Interest Expense", 1),
            (13, "EBT", 1), (14, "Tax Provision", 1), (15, "Net Income", 0),
            (17, "EBITDA Margin %", 1), (18, "Net Margin %", 1)
        ]:
            fmt.apply_label_cell(row_num, 2, label, indent=indent)

        fmt.apply_header_row(6, 2, dc + n - 1, "REVENUE")
        fmt.apply_header_row(16, 2, dc + n - 1, "MARGIN ANALYSIS")
        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_debt_schedule(self):
        ws, fmt = self.add_sheet("DEBT_SCHEDULE", tab_color="C00000")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"
        su = "SOURCES_USES"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "DEBT SCHEDULE", "Debt Paydown Waterfall & Cash Sweep")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"Year {i+1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        # SENIOR DEBT
        fmt.apply_header_row(5, 2, dc + n - 1, "SENIOR DEBT")
        for row_num, label in [(6, "Opening Balance"), (7, "Opening Balance"), (8, "Scheduled Amortization"),
                                (9, "Cash Sweep"), (10, "Total Repayment"), (11, "Closing Balance")]:
            fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                open_bal = f"={su}!G6"
            else:
                open_bal = f"={prev_cl}11"

            ws.cell(row=7, column=col, value=open_bal).number_format = NumFormats.USD_MILLIONS
            ws.cell(row=8, column=col, value=f"=-{cl}7*{asm}!C18").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=9, column=col, value=f"=-MAX(CASH_FLOW!{cl}10*{asm}!C19,0)").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=10, column=col, value=f"={cl}8+{cl}9").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=11, column=col, value=f"=MAX({cl}7+{cl}10,0)").number_format = NumFormats.USD_MILLIONS

            for row_num in [7, 8, 9, 10, 11]:
                c = ws.cell(row=row_num, column=col)
                c.font = Fonts.body()
                c.fill = Fills.white()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

            bold_rows = [11]
            for r in bold_rows:
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Calibri", size=11, bold=True)
                c.fill = Fills.subheader()
                c.border = Borders.thick_bottom()

        # SUB DEBT
        fmt.apply_header_row(13, 2, dc + n - 1, "SUBORDINATED DEBT")
        for row_num, label in [(14, "Opening Balance"), (15, "PIK Interest Accrual"),
                                (16, "Cash Interest"), (17, "Closing Balance")]:
            fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                sub_open = f"={su}!G7"
            else:
                sub_open = f"={prev_cl}17"

            ws.cell(row=14, column=col, value=sub_open).number_format = NumFormats.USD_MILLIONS
            ws.cell(row=15, column=col, value=f"=0").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=16, column=col, value=f"={cl}14*{asm}!C16").number_format = NumFormats.USD_MILLIONS
            ws.cell(row=17, column=col, value=f"={cl}14+{cl}15").number_format = NumFormats.USD_MILLIONS

            for row_num in [14, 15, 16, 17]:
                c = ws.cell(row=row_num, column=col)
                c.font = Fonts.body()
                c.fill = Fills.white()
                c.border = Borders.bottom_only()
                c.alignment = Alignment(horizontal="right")

        # Total Debt
        fmt.apply_header_row(19, 2, dc + n - 1, "TOTAL DEBT SUMMARY")
        for row_num, label in [(20, "Total Debt (Closing)"), (21, "Leverage (x EBITDA)")]:
            fmt.apply_label_cell(row_num, 2, label, indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)

            td = ws.cell(row=20, column=col, value=f"={cl}11+{cl}17")
            td.number_format = NumFormats.USD_MILLIONS
            td.font = Font(name="Calibri", size=11, bold=True)
            td.fill = Fills.subheader()
            td.border = Borders.thick_bottom()
            td.alignment = Alignment(horizontal="right")

            lev = ws.cell(row=21, column=col, value=f"={cl}20/INCOME_STATEMENT!{cl}9")
            lev.number_format = NumFormats.MULTIPLE
            lev.font = Fonts.body()
            lev.border = Borders.bottom_only()
            lev.alignment = Alignment(horizontal="right")

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_cash_flow(self):
        ws, fmt = self.add_sheet("CASH_FLOW", tab_color="375623")
        n = self.years
        dc = self.dc
        asm = "ASSUMPTIONS"
        is_sheet = "INCOME_STATEMENT"

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "CASH FLOW STATEMENT", "Free Cash Flow & Debt Service Waterfall")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"Year {i+1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        cf_rows = [
            (5,  "OPERATING CASH FLOW",    None,                                                True),
            (6,  "EBITDA",                 f"={is_sheet}!{{col}}9",                            False),
            (7,  "Less: Capex",            f"=-{is_sheet}!{{col}}7*{asm}!C25",                False),
            (8,  "Less: Change in NWC",    f"=-{is_sheet}!{{col}}7*{asm}!C26",                False),
            (9,  "Less: Cash Taxes",       f"=-{is_sheet}!{{col}}13*{asm}!C27",               False),
            (10, "Cash Available for Debt Service", f"={{col}}6+{{col}}7+{{col}}8+{{col}}9",  True),
            (11, "Less: Cash Interest (Senior)",    f"=-DEBT_SCHEDULE!{{col}}7*{asm}!C14",   False),
            (12, "Less: Cash Interest (Sub)",       f"=-DEBT_SCHEDULE!{{col}}16",             False),
            (13, "Net Cash After Interest",         f"={{col}}10+{{col}}11+{{col}}12",         True),
            (14, "Less: Mandatory Amortization",    f"=DEBT_SCHEDULE!{{col}}8",               False),
            (15, "Free Cash Flow (Pre-Sweep)",      f"={{col}}13+{{col}}14",                   True),
        ]

        for row_num, label, formula_tpl, is_total in cf_rows:
            if is_total:
                fmt.apply_header_row(row_num, 2, dc + n - 1, label)
            else:
                fmt.apply_label_cell(row_num, 2, label, indent=1)

            if formula_tpl:
                for i in range(n):
                    col = dc + i
                    cl = get_column_letter(col)
                    formula = formula_tpl.replace("{col}", cl)
                    c = ws.cell(row=row_num, column=col, value=formula)
                    c.number_format = NumFormats.USD_MILLIONS
                    c.alignment = Alignment(horizontal="right")
                    if is_total and row_num != 5:
                        c.font = Font(name="Calibri", size=11, bold=True)
                        c.fill = Fills.subheader()
                        c.border = Borders.thick_bottom()
                    else:
                        c.font = Fonts.body()
                        c.fill = Fills.white()
                        c.border = Borders.bottom_only()

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    def _build_returns(self):
        ws, fmt = self.add_sheet("RETURNS", tab_color=Colors.ACCENT_GOLD)
        n = self.years
        asm = "ASSUMPTIONS"
        is_sheet = "INCOME_STATEMENT"
        ds = "DEBT_SCHEDULE"
        su = "SOURCES_USES"

        fmt.set_column_widths({"A": 3, "B": 40, "C": 22, "D": 22, "E": 22})
        fmt.apply_sheet_title(2, 2, "RETURNS ANALYSIS", "IRR / MOIC — Sponsor Returns")

        # Entry
        fmt.apply_header_row(5, 2, 4, "ENTRY")
        entry_rows = [
            (6,  "Entry Enterprise Value",    f"={asm}!C7",          NumFormats.USD_MILLIONS),
            (7,  "Entry EBITDA",              f"={asm}!C6",           NumFormats.USD_MILLIONS),
            (8,  "Entry Multiple (x)",        f"={asm}!C8",           NumFormats.MULTIPLE),
            (9,  "Total Debt at Entry",       f"={su}!C14",           NumFormats.USD_MILLIONS),
            (10, "Sponsor Equity Invested",   f"={su}!G8",            NumFormats.USD_MILLIONS),
        ]
        for row_num, label, formula, fmt_code in entry_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.number_format = fmt_code
            c.font = Fonts.body()
            c.border = Borders.bottom_only()
            c.alignment = Alignment(horizontal="right")

        # Exit
        last_cl = get_column_letter(self.dc + n - 1)
        fmt.apply_header_row(12, 2, 4, "EXIT")
        exit_rows = [
            (13, "Exit Year EBITDA",        f"={is_sheet}!{last_cl}9",  NumFormats.USD_MILLIONS),
            (14, "Exit Multiple (x)",       f"={asm}!C30",              NumFormats.MULTIPLE),
            (15, "Exit Enterprise Value",   f"=C13*C14",                NumFormats.USD_MILLIONS),
            (16, "Less: Remaining Debt",    f"=-{ds}!{last_cl}20",      NumFormats.USD_MILLIONS),
            (17, "Exit Equity Value",       f"=C15+C16",                NumFormats.USD_MILLIONS),
        ]
        for row_num, label, formula, fmt_code in exit_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.number_format = fmt_code
            c.font = Fonts.body() if row_num != 17 else Font(name="Calibri", size=11, bold=True)
            c.fill = Fills.white() if row_num != 17 else Fills.subheader()
            c.border = Borders.bottom_only() if row_num != 17 else Borders.thick_bottom()
            c.alignment = Alignment(horizontal="right")

        # Returns
        fmt.apply_header_row(19, 2, 4, "SPONSOR RETURNS")
        eq_invested = f"={su}!G8"
        eq_exit = "=C17"

        ws.cell(row=20, column=2, value="MOIC")
        moic = ws.cell(row=20, column=3, value=f"=C17/{su}!G8")
        moic.number_format = NumFormats.MULTIPLE
        moic.font = Font(name="Calibri", size=20, bold=True, color=Colors.INST_BLUE)
        moic.alignment = Alignment(horizontal="center")
        ws.row_dimensions[20].height = 40

        ws.cell(row=21, column=2, value="IRR")
        # Build cash flows: [-equity, 0, 0, ..., exit_equity]
        cf_series = f"-{su}!G8"
        mid_zeros = ",".join(["0"] * (n - 1))
        irr_formula = f"=IRR({{{cf_series},{mid_zeros},C17}})"
        irr = ws.cell(row=21, column=3, value=irr_formula)
        irr.number_format = NumFormats.PERCENT_ONE
        irr.font = Font(name="Calibri", size=20, bold=True, color=Colors.INST_BLUE)
        irr.alignment = Alignment(horizontal="center")
        ws.row_dimensions[21].height = 40

        # Sensitivity: Exit Multiple vs Entry Multiple
        fmt.apply_header_row(23, 2, 8, "IRR SENSITIVITY — EXIT MULTIPLE VS LEVERAGE")
        ws.cell(row=24, column=2, value="Entry Lev ↓ / Exit Mult →")
        exit_mults = [7.0, 8.0, 9.0, 10.0, 11.0]
        entry_levs = [3.0, 3.5, 4.0, 4.5, 5.0]

        for j, em in enumerate(exit_mults):
            c = ws.cell(row=24, column=3 + j, value=em)
            c.number_format = NumFormats.MULTIPLE
            c.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            c.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            c.alignment = Alignment(horizontal="center")

        for i, el in enumerate(entry_levs):
            row = 25 + i
            lbl = ws.cell(row=row, column=2, value=el)
            lbl.number_format = NumFormats.MULTIPLE
            lbl.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            lbl.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            lbl.alignment = Alignment(horizontal="center")

            for j, em in enumerate(exit_mults):
                # Simplified IRR sensitivity approximation
                approx_exit_eq = f"=({is_sheet}!{last_cl}9*{em}-{ds}!{last_cl}20)"
                irr_val = f"=IFERROR(({is_sheet}!{last_cl}9*{em}-{ds}!{last_cl}20)/{su}!G8,\"N/A\")"
                c = ws.cell(row=row, column=3 + j, value=irr_val)
                c.number_format = NumFormats.MULTIPLE
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center")
                c.fill = Fills.sensitivity_mid()

    def _build_dashboard(self):
        ws, fmt = self.add_sheet("DASHBOARD", tab_color=Colors.ACCENT_GOLD)
        n = self.years
        su = "SOURCES_USES"
        ret = "RETURNS"
        ds = "DEBT_SCHEDULE"
        asm = "ASSUMPTIONS"

        ws.sheet_view.showGridLines = False
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["A"].width = 2

        for row in range(1, 3):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
        for row in range(3, 4):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)

        ws.merge_cells("B2:O2")
        title = ws["B2"]
        title.value = f"LBO MODEL DASHBOARD — {self.a.get('company_name','Target Co.')}"
        title.font = Font(name="Calibri", size=16, bold=True, color=Colors.WHITE)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 30

        # KPI Cards
        kpis = [
            ("B", "IRR",           f"={ret}!C21",  NumFormats.PERCENT_ONE),
            ("E", "MOIC",          f"={ret}!C20",  NumFormats.MULTIPLE),
            ("H", "Exit Eq Value", f"={ret}!C17",  NumFormats.USD_MILLIONS),
            ("K", "Debt Paydown",  f"={su}!C14-{ds}!{get_column_letter(self.dc+n-1)}20", NumFormats.USD_MILLIONS),
            ("N", "Hold Period",   f"={asm}!C10",  NumFormats.INTEGER),
        ]

        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 38

        for col_letter, label, formula, fmt_code in kpis:
            lbl = ws[f"{col_letter}5"]
            lbl.value = label
            lbl.font = Fonts.kpi_label()
            lbl.fill = Fills.kpi_card()
            lbl.alignment = Alignment(horizontal="center", vertical="center")
            lbl.border = Border(
                top=Side(style="medium", color=Colors.INST_BLUE),
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE)
            )
            val = ws[f"{col_letter}6"]
            val.value = formula
            val.font = Fonts.kpi_value()
            val.number_format = fmt_code
            val.fill = Fills.kpi_card()
            val.alignment = Alignment(horizontal="center", vertical="center")
            val.border = Border(
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE),
                bottom=Side(style="medium", color=Colors.INST_BLUE)
            )

        # Chart data
        chart_row = 50
        for i in range(n):
            ws.cell(row=chart_row, column=3 + i, value=f"Year {i+1}")
            ws.cell(row=chart_row + 1, column=3 + i, value=f"=DEBT_SCHEDULE!{get_column_letter(self.dc+i)}20")
            ws.cell(row=chart_row + 2, column=3 + i, value=f"=INCOME_STATEMENT!{get_column_letter(self.dc+i)}9")

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Debt Paydown vs EBITDA Growth"
        chart1.style = 10
        chart1.width = 14
        chart1.height = 10

        debt_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_row + 1)
        ebitda_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_row + 2)
        cats = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_row)
        chart1.add_data(debt_ref)
        chart1.add_data(ebitda_ref)
        chart1.set_categories(cats)
        from openpyxl.chart.series import SeriesLabel
        chart1.series[0].title = SeriesLabel(v="Total Debt")
        chart1.series[1].title = SeriesLabel(v="EBITDA")
        ws.add_chart(chart1, "B8")

        fmt.hide_gridlines()
