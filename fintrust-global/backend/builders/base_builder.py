"""
Base Builder — Shared Excel utilities for all model types
Financial Modeling Agent
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from formatting.institutional import Colors, Fonts, Fills, Borders, NumFormats, ColWidths, Formatter
import io


class BaseBuilder:

    def __init__(self, assumptions: dict):
        self.assumptions = assumptions
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

    def add_sheet(self, name: str, tab_color: str = None):
        ws = self.wb.create_sheet(title=name)
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        fmt = Formatter(ws)
        fmt.hide_gridlines()
        fmt.set_zoom(85)
        return ws, fmt

    def build_cover(self, model_type: str):
        ws, fmt = self.add_sheet("COVER", tab_color=Colors.DARK_NAVY)
        a = self.assumptions

        # Full navy background
        for row in range(1, 45):
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.fill = Fills.cover_navy()

        # Accent bar — col 1
        for row in range(1, 45):
            cell = ws.cell(row=row, column=1)
            cell.fill = PatternFill("solid", fgColor=Colors.INST_BLUE)

        # Title block
        ws.merge_cells("C8:N8")
        t = ws["C8"]
        t.value = a.get("company_name", "Company Name")
        t.font = Fonts.cover_main()
        t.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("C9:N9")
        s = ws["C9"]
        s.value = f"{model_type} Financial Model"
        s.font = Font(name="Calibri", size=18, color="BDD7EE")
        s.alignment = Alignment(horizontal="left", vertical="center")

        # Divider line
        ws.merge_cells("C11:N11")
        d = ws["C11"]
        d.value = ""
        d.border = Border(bottom=Side(style="medium", color=Colors.ACCENT_GOLD))

        # Meta info
        meta = [
            ("Prepared by:", a.get("analyst_name", "Financial Analyst")),
            ("Date:", a.get("model_date", "2024")),
            ("Currency:", a.get("currency", "USD in Millions")),
            ("Projection Period:", f"{a.get('projection_years', 5)} Years"),
        ]
        for i, (label, value) in enumerate(meta):
            r = 13 + i
            lbl = ws.cell(row=r, column=3, value=label)
            lbl.font = Font(name="Calibri", size=11, color="BDD7EE")
            val = ws.cell(row=r, column=5, value=value)
            val.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)

        # Confidentiality notice
        ws.merge_cells("C30:N30")
        conf = ws["C30"]
        conf.value = "CONFIDENTIAL — FOR INTERNAL USE ONLY"
        conf.font = Font(name="Calibri", size=9, color="808080", italic=True)
        conf.alignment = Alignment(horizontal="left")

        # Sheet legend
        ws.merge_cells("C32:N32")
        leg_title = ws["C32"]
        leg_title.value = "COLOR LEGEND"
        leg_title.font = Font(name="Calibri", size=10, bold=True, color="BDD7EE")

        legend = [
            (Colors.INPUT_BG,    Colors.INPUT_BLUE,    "Blue — User Input / Assumptions"),
            (Colors.WHITE,       Colors.FORMULA_BLACK,  "Black — Formula / Calculated"),
            (Colors.LIGHT_GREY,  Colors.EXTERNAL_GREEN, "Green — External Link / Reference"),
            (Colors.HEADER_BG,   Colors.FORMULA_BLACK,  "Grey — Section Header"),
        ]
        for i, (bg, fg, text) in enumerate(legend):
            r = 33 + i
            swatch = ws.cell(row=r, column=3, value="  ")
            swatch.fill = PatternFill("solid", fgColor=bg)
            swatch.border = Borders.thin()
            label = ws.cell(row=r, column=4, value=text)
            label.font = Font(name="Calibri", size=10, color=fg)
            label.fill = Fills.cover_navy()

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 3
        for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
            ws.column_dimensions[col_letter].width = 14
        ws.row_dimensions[8].height = 50
        ws.row_dimensions[9].height = 30
        fmt.hide_gridlines()

    def save_to_buffer(self) -> bytes:
        """Save workbook to bytes buffer for download"""
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _write_section_spacer(self, ws, row, col):
        """Write empty spacer row"""
        ws.row_dimensions[row].height = 6

    def _col(self, n):
        """Return Excel column letter for column number n"""
        return get_column_letter(n)
