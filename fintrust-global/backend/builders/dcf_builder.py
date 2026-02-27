"""
DCF Model Builder — 7 Sheet Institutional Model
Sheets: COVER, ASSUMPTIONS, INCOME_STATEMENT, FCF_BRIDGE, WACC, VALUATION, DASHBOARD
"""

from builders.base_builder import BaseBuilder
from formatting.institutional import (
    Colors, Fonts, Fills, Borders, NumFormats, ColWidths, Formatter
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint


class DCFBuilder(BaseBuilder):

    def __init__(self, assumptions: dict):
        super().__init__(assumptions)
        self.a = assumptions
        self.years = int(self.a.get("projection_years", 5))
        self.base_year = int(self.a.get("base_year", 2024))
        self.data_col_start = 3  # Column C = first data column
        self.label_col = 2       # Column B = row labels

    def build(self):
        self.build_cover("DCF")
        self._build_assumptions()
        self._build_income_statement()
        self._build_fcf_bridge()
        self._build_wacc()
        self._build_valuation()
        self._build_dashboard()
        return self.save_to_buffer()

    # ─── SHEET 2: ASSUMPTIONS ───────────────────────────────
    def _build_assumptions(self):
        ws, fmt = self.add_sheet("ASSUMPTIONS", tab_color="2E75B6")
        a = self.a

        fmt.set_column_widths({
            "A": 3, "B": 38, "C": 20, "D": 20, "E": 20
        })
        fmt.freeze_panes("A1")
        fmt.apply_sheet_title(2, 2, "MODEL ASSUMPTIONS", "All blue cells are editable inputs")
        fmt.apply_units_label(3, 2, "$ in Millions unless stated")

        # ── COMPANY & MODEL INFO ──
        fmt.apply_header_row(5, 2, 5, "COMPANY & MODEL INFORMATION")
        labels = ["Company Name", "Model Date", "Analyst", "Base Year", "Projection Years", "Currency"]
        defaults = [
            a.get("company_name", "Target Company"),
            a.get("model_date", "2024"),
            a.get("analyst_name", "Financial Analyst"),
            str(self.base_year),
            str(self.years),
            a.get("currency", "USD")
        ]
        for i, (lbl, val) in enumerate(zip(labels, defaults)):
            fmt.apply_label_cell(6 + i, 2, lbl)
            fmt.apply_input_cell(6 + i, 3, val)

        # ── REVENUE ASSUMPTIONS ──
        fmt.apply_header_row(13, 2, 5, "REVENUE ASSUMPTIONS")
        rev_labels = [
            ("Base Year Revenue ($M)",   "base_revenue",    100.0,  NumFormats.USD_MILLIONS),
            ("Year 1 Growth Rate",       "rev_growth_y1",   0.10,   NumFormats.PERCENT_ONE),
            ("Year 2 Growth Rate",       "rev_growth_y2",   0.09,   NumFormats.PERCENT_ONE),
            ("Year 3 Growth Rate",       "rev_growth_y3",   0.08,   NumFormats.PERCENT_ONE),
            ("Year 4 Growth Rate",       "rev_growth_y4",   0.07,   NumFormats.PERCENT_ONE),
            ("Year 5 Growth Rate",       "rev_growth_y5",   0.06,   NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(rev_labels):
            fmt.apply_label_cell(14 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(14 + i, 3, val, fmt_code)

        # ── MARGIN ASSUMPTIONS ──
        fmt.apply_header_row(21, 2, 5, "MARGIN ASSUMPTIONS")
        margin_labels = [
            ("Gross Margin %",       "gross_margin",   0.60,  NumFormats.PERCENT_ONE),
            ("EBITDA Margin %",      "ebitda_margin",  0.25,  NumFormats.PERCENT_ONE),
            ("D&A as % of Revenue",  "da_pct",         0.05,  NumFormats.PERCENT_ONE),
            ("Tax Rate",             "tax_rate",       0.25,  NumFormats.PERCENT_ONE),
            ("Capex as % of Revenue","capex_pct",      0.06,  NumFormats.PERCENT_ONE),
            ("Change in NWC as % Rev","nwc_pct",       0.02,  NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(margin_labels):
            fmt.apply_label_cell(22 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(22 + i, 3, val, fmt_code)

        # ── WACC INPUTS ──
        fmt.apply_header_row(29, 2, 5, "WACC ASSUMPTIONS")
        wacc_labels = [
            ("Risk-Free Rate",          "risk_free_rate",  0.045, NumFormats.PERCENT_ONE),
            ("Equity Risk Premium",     "erp",             0.055, NumFormats.PERCENT_ONE),
            ("Beta (Levered)",          "beta",            1.10,  "0.00"),
            ("Cost of Debt (Pre-Tax)",  "cost_of_debt",    0.06,  NumFormats.PERCENT_ONE),
            ("Debt / Total Capital",    "debt_weight",     0.30,  NumFormats.PERCENT_ONE),
            ("Equity / Total Capital",  "equity_weight",   0.70,  NumFormats.PERCENT_ONE),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(wacc_labels):
            fmt.apply_label_cell(30 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(30 + i, 3, val, fmt_code)

        # ── TERMINAL VALUE ──
        fmt.apply_header_row(37, 2, 5, "TERMINAL VALUE")
        tv_labels = [
            ("Terminal Growth Rate",    "terminal_growth",  0.025, NumFormats.PERCENT_ONE),
            ("Exit EV/EBITDA Multiple", "exit_multiple",    10.0,  NumFormats.MULTIPLE),
            ("Shares Outstanding (M)",  "shares_out",       50.0,  NumFormats.INTEGER),
            ("Net Debt ($M)",           "net_debt",         20.0,  NumFormats.USD_MILLIONS),
        ]
        for i, (lbl, key, default, fmt_code) in enumerate(tv_labels):
            fmt.apply_label_cell(38 + i, 2, lbl)
            val = float(a.get(key, default))
            fmt.apply_input_cell(38 + i, 3, val, fmt_code)

    # ─── SHEET 3: INCOME STATEMENT ──────────────────────────
    def _build_income_statement(self):
        ws, fmt = self.add_sheet("INCOME_STATEMENT", tab_color="375623")
        n = self.years
        dc = self.data_col_start

        fmt.set_column_widths({"A": 3, "B": 38, "C": 3})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "INCOME STATEMENT", f"Projected {n}-Year P&L")
        fmt.apply_units_label(3, 2, "$ in Millions")
        fmt.apply_header_row(4, 2, dc + n, "")

        # Timeline headers
        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"FY{self.base_year + i + 1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row definitions: (label, row_num, formula_template, num_format, indent, is_total)
        rows = [
            # Section: Revenue
            ("REVENUE", 6, None, None, 0, True),
            ("Net Revenue", 7, None, NumFormats.USD_MILLIONS, 1, False),

            # Section: Cost Structure
            ("COST STRUCTURE", 9, None, None, 0, True),
            ("Cost of Goods Sold", 10, None, NumFormats.USD_MILLIONS, 1, False),
            ("Gross Profit", 11, None, NumFormats.USD_MILLIONS, 1, False),

            # EBITDA
            ("EBITDA", 13, None, None, 0, True),
            ("Operating Expenses", 14, None, NumFormats.USD_MILLIONS, 1, False),
            ("EBITDA", 15, None, NumFormats.USD_MILLIONS, 1, False),
            ("Depreciation & Amortization", 16, None, NumFormats.USD_MILLIONS, 1, False),
            ("EBIT", 17, None, NumFormats.USD_MILLIONS, 1, False),

            # Below EBIT
            ("BELOW THE LINE", 19, None, None, 0, True),
            ("Interest Expense", 20, None, NumFormats.USD_MILLIONS, 1, False),
            ("EBT", 21, None, NumFormats.USD_MILLIONS, 1, False),
            ("Tax Provision", 22, None, NumFormats.USD_MILLIONS, 1, False),
            ("Net Income", 23, None, NumFormats.USD_MILLIONS, 1, False),

            # Margins
            ("MARGIN ANALYSIS", 25, None, None, 0, True),
            ("Gross Margin %", 26, None, NumFormats.PERCENT_ONE, 1, False),
            ("EBITDA Margin %", 27, None, NumFormats.PERCENT_ONE, 1, False),
            ("EBIT Margin %", 28, None, NumFormats.PERCENT_ONE, 1, False),
            ("Net Margin %", 29, None, NumFormats.PERCENT_ONE, 1, False),
            ("Revenue Growth %", 30, None, NumFormats.PERCENT_ONE, 1, False),
        ]

        for label, row, _, fmt_code, indent, is_total in rows:
            if is_total:
                fmt.apply_header_row(row, 2, dc + n - 1, label)
            else:
                fmt.apply_label_cell(row, 2, label, indent=indent)

        # Populate formulas per year
        asm = "ASSUMPTIONS"
        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if i == 0:
                rev = f"={asm}!C14*(1+{asm}!C{15+i})"
            else:
                rev = f"={get_column_letter(col-1)}7*(1+{asm}!C{15+i})"

            formulas = {
                7:  rev,                                       # Revenue
                10: f"=-{cl}7*{asm}!C22",                     # COGS
                11: f"={cl}7+{cl}10",                         # Gross Profit
                14: f"=-{cl}7*(1-{asm}!C23)",                 # OpEx (implied)
                15: f"={cl}7*{asm}!C23",                      # EBITDA
                16: f"=-{cl}7*{asm}!C24",                     # D&A
                17: f"={cl}15+{cl}16",                        # EBIT
                20: f"=0",                                     # Interest (0 for unlevered DCF)
                21: f"={cl}17+{cl}20",                        # EBT
                22: f"=-{cl}21*{asm}!C25",                    # Tax
                23: f"={cl}21+{cl}22",                        # Net Income
                26: f"={cl}11/{cl}7",                         # Gross Margin
                27: f"={cl}15/{cl}7",                         # EBITDA Margin
                28: f"={cl}17/{cl}7",                         # EBIT Margin
                29: f"={cl}23/{cl}7",                         # Net Margin
                30: f"=IF({col}>3,{cl}7/{prev_cl}7-1,\"\")", # Rev Growth
            }

            formats = {
                7: NumFormats.USD_MILLIONS, 10: NumFormats.USD_MILLIONS,
                11: NumFormats.USD_MILLIONS, 14: NumFormats.USD_MILLIONS,
                15: NumFormats.USD_MILLIONS, 16: NumFormats.USD_MILLIONS,
                17: NumFormats.USD_MILLIONS, 20: NumFormats.USD_MILLIONS,
                21: NumFormats.USD_MILLIONS, 22: NumFormats.USD_MILLIONS,
                23: NumFormats.USD_MILLIONS,
                26: NumFormats.PERCENT_ONE, 27: NumFormats.PERCENT_ONE,
                28: NumFormats.PERCENT_ONE, 29: NumFormats.PERCENT_ONE,
                30: NumFormats.PERCENT_ONE,
            }

            for row_num, formula in formulas.items():
                cell = ws.cell(row=row_num, column=col, value=formula)
                cell.font = Fonts.body()
                cell.fill = Fills.white()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Borders.bottom_only()
                if row_num in formats:
                    cell.number_format = formats[row_num]

        # Bold totals
        for row_num in [11, 15, 17, 21, 23]:
            for i in range(n):
                col = dc + i
                c = ws.cell(row=row_num, column=col)
                c.font = Font(name="Calibri", size=11, bold=True)
                c.fill = Fills.subheader()
                c.border = Borders.thick_bottom()

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    # ─── SHEET 4: FCF BRIDGE ────────────────────────────────
    def _build_fcf_bridge(self):
        ws, fmt = self.add_sheet("FCF_BRIDGE", tab_color="375623")
        n = self.years
        dc = self.data_col_start

        fmt.set_column_widths({"A": 3, "B": 38})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR

        fmt.apply_sheet_title(2, 2, "FREE CASH FLOW BRIDGE", "UFCF Derivation from NOPAT")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"FY{self.base_year + i + 1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        fmt.apply_header_row(5, 2, dc + n - 1, "UNLEVERED FREE CASH FLOW BRIDGE")

        is_sheet = "INCOME_STATEMENT"
        asm_sheet = "ASSUMPTIONS"
        fcf_rows = [
            (6,  "EBIT",                        f"={is_sheet}!{{col}}17",       NumFormats.USD_MILLIONS),
            (7,  "Less: Taxes on EBIT",          f"=-{is_sheet}!{{col}}17*{asm_sheet}!C25", NumFormats.USD_MILLIONS),
            (8,  "NOPAT",                        f"={is_sheet}!{{col}}17*(1-{asm_sheet}!C25)", NumFormats.USD_MILLIONS),
            (9,  "Add: D&A",                     f"=-{is_sheet}!{{col}}16",      NumFormats.USD_MILLIONS),
            (10, "Less: Capital Expenditures",   f"=-{is_sheet}!{{col}}7*{asm_sheet}!C26", NumFormats.USD_MILLIONS),
            (11, "Less: Change in Net Working Capital", f"=-{is_sheet}!{{col}}7*{asm_sheet}!C27", NumFormats.USD_MILLIONS),
            (12, "Unlevered Free Cash Flow",     f"={{col}}8+{{col}}9+{{col}}10+{{col}}11", NumFormats.USD_MILLIONS),
            (14, "FCF Margin %",                 f"={{col}}12/{is_sheet}!{{col}}7", NumFormats.PERCENT_ONE),
        ]

        for row_num, label, formula_tpl, fmt_code in fcf_rows:
            is_total = row_num in [8, 12]
            if is_total:
                fmt.apply_total_row(row_num, 2, dc + n - 1, label, fmt_code)
            else:
                fmt.apply_label_cell(row_num, 2, label, indent=1)

            for i in range(n):
                col = dc + i
                cl = get_column_letter(col)
                formula = formula_tpl.replace("{col}", cl)
                c = ws.cell(row=row_num, column=col, value=formula)
                c.font = Fonts.body()
                c.fill = Fills.white() if not is_total else Fills.subheader()
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.border = Borders.bottom_only()
                c.number_format = fmt_code

        fmt.freeze_panes(f"{get_column_letter(dc)}5")

    # ─── SHEET 5: WACC ──────────────────────────────────────
    def _build_wacc(self):
        ws, fmt = self.add_sheet("WACC", tab_color="7030A0")
        asm = "ASSUMPTIONS"

        fmt.set_column_widths({"A": 3, "B": 40, "C": 22, "D": 22})
        fmt.apply_sheet_title(2, 2, "WACC CALCULATION", "Weighted Average Cost of Capital")

        fmt.apply_header_row(5, 2, 4, "COST OF EQUITY — CAPM")
        capm_rows = [
            (6,  "Risk-Free Rate",       f"={asm}!C30", NumFormats.PERCENT_ONE),
            (7,  "Equity Risk Premium",  f"={asm}!C31", NumFormats.PERCENT_ONE),
            (8,  "Levered Beta",         f"={asm}!C32", "0.00"),
            (9,  "Cost of Equity (Ke)",  f"={asm}!C30+{asm}!C32*{asm}!C31", NumFormats.PERCENT_ONE),
        ]
        for row_num, label, formula, fmt_code in capm_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.font = Fonts.body() if row_num != 9 else Font(name="Calibri", size=11, bold=True)
            c.fill = Fills.white() if row_num != 9 else Fills.subheader()
            c.alignment = Alignment(horizontal="right")
            c.border = Borders.bottom_only()
            c.number_format = fmt_code

        fmt.apply_header_row(11, 2, 4, "COST OF DEBT — AFTER-TAX")
        debt_rows = [
            (12, "Pre-Tax Cost of Debt",  f"={asm}!C33", NumFormats.PERCENT_ONE),
            (13, "Tax Rate",              f"={asm}!C25", NumFormats.PERCENT_ONE),
            (14, "After-Tax Cost of Debt (Kd)", f"={asm}!C33*(1-{asm}!C25)", NumFormats.PERCENT_ONE),
        ]
        for row_num, label, formula, fmt_code in debt_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.font = Fonts.body() if row_num != 14 else Font(name="Calibri", size=11, bold=True)
            c.fill = Fills.white() if row_num != 14 else Fills.subheader()
            c.alignment = Alignment(horizontal="right")
            c.border = Borders.bottom_only()
            c.number_format = fmt_code

        fmt.apply_header_row(16, 2, 4, "CAPITAL STRUCTURE")
        cap_rows = [
            (17, "Equity Weight",  f"={asm}!C35", NumFormats.PERCENT_ONE),
            (18, "Debt Weight",    f"={asm}!C34", NumFormats.PERCENT_ONE),
            (19, "Check (= 100%)", f"={asm}!C35+{asm}!C34", NumFormats.PERCENT_ONE),
        ]
        for row_num, label, formula, fmt_code in cap_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.font = Fonts.body()
            c.fill = Fills.white()
            c.alignment = Alignment(horizontal="right")
            c.border = Borders.bottom_only()
            c.number_format = fmt_code

        fmt.apply_header_row(21, 2, 4, "BLENDED WACC")
        ws.cell(row=22, column=2, value="WACC").font = Font(name="Calibri", size=14, bold=True, color=Colors.DARK_NAVY)
        wacc_cell = ws.cell(
            row=22, column=3,
            value=f"={asm}!C35*({asm}!C30+{asm}!C32*{asm}!C31)+{asm}!C34*{asm}!C33*(1-{asm}!C25)"
        )
        wacc_cell.font = Font(name="Calibri", size=20, bold=True, color=Colors.INST_BLUE)
        wacc_cell.number_format = NumFormats.PERCENT_ONE
        wacc_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[22].height = 40

    # ─── SHEET 6: VALUATION ─────────────────────────────────
    def _build_valuation(self):
        ws, fmt = self.add_sheet("VALUATION", tab_color="C00000")
        n = self.years
        dc = self.data_col_start
        asm = "ASSUMPTIONS"
        wacc_cell = "WACC!C22"
        fcf_sheet = "FCF_BRIDGE"

        fmt.set_column_widths({"A": 3, "B": 40})
        for i in range(n):
            ws.column_dimensions[get_column_letter(dc + i)].width = ColWidths.YEAR
        ws.column_dimensions[get_column_letter(dc + n)].width = 20

        fmt.apply_sheet_title(2, 2, "DCF VALUATION", "Enterprise Value to Equity Value Bridge")
        fmt.apply_units_label(3, 2, "$ in Millions")

        for i in range(n):
            col = dc + i
            cell = ws.cell(row=4, column=col, value=f"FY{self.base_year + i + 1}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center")

        fmt.apply_header_row(5, 2, dc + n - 1, "DCF — PRESENT VALUE OF FREE CASH FLOWS")
        fmt.apply_label_cell(6, 2, "Unlevered Free Cash Flow", indent=1)
        fmt.apply_label_cell(7, 2, "Discount Factor", indent=1)
        fmt.apply_label_cell(8, 2, "PV of Free Cash Flow", indent=1)

        for i in range(n):
            col = dc + i
            cl = get_column_letter(col)
            period = i + 1

            fcf = ws.cell(row=6, column=col, value=f"={fcf_sheet}!{cl}12")
            fcf.number_format = NumFormats.USD_MILLIONS
            fcf.font = Fonts.external()
            fcf.border = Borders.bottom_only()
            fcf.alignment = Alignment(horizontal="right")

            disc = ws.cell(row=7, column=col, value=f"=1/(1+{wacc_cell})^{period}")
            disc.number_format = "0.0000"
            disc.font = Fonts.body()
            disc.border = Borders.bottom_only()
            disc.alignment = Alignment(horizontal="right")

            pv = ws.cell(row=8, column=col, value=f"={cl}6*{cl}7")
            pv.number_format = NumFormats.USD_MILLIONS
            pv.font = Font(name="Calibri", size=11, bold=True)
            pv.fill = Fills.subheader()
            pv.border = Borders.thick_bottom()
            pv.alignment = Alignment(horizontal="right")

        # Sum of PV FCFs
        pv_cols = "+".join([f"{get_column_letter(dc+i)}8" for i in range(n)])
        fmt.apply_header_row(10, 2, dc + n - 1, "TERMINAL VALUE")
        fmt.apply_label_cell(11, 2, "Terminal Year UFCF", indent=1)
        fmt.apply_label_cell(12, 2, "Terminal Growth Rate", indent=1)
        fmt.apply_label_cell(13, 2, "WACC", indent=1)
        fmt.apply_label_cell(14, 2, "Gordon Growth Terminal Value (Undiscounted)", indent=1)
        fmt.apply_label_cell(15, 2, "Exit Multiple Terminal Value (Undiscounted)", indent=1)
        fmt.apply_label_cell(16, 2, "PV of Terminal Value (Gordon Growth)", indent=1)

        last_fcf_col = get_column_letter(dc + n - 1)

        for row_num, val, fmt_code in [
            (11, f"={fcf_sheet}!{last_fcf_col}12", NumFormats.USD_MILLIONS),
            (12, f"={asm}!C38",  NumFormats.PERCENT_ONE),
            (13, f"={wacc_cell}", NumFormats.PERCENT_ONE),
            (14, f"=C11*(1+C12)/(C13-C12)", NumFormats.USD_MILLIONS),
            (15, f"=INCOME_STATEMENT!{last_fcf_col}15*{asm}!C39", NumFormats.USD_MILLIONS),
            (16, f"=C14/(1+{wacc_cell})^{n}", NumFormats.USD_MILLIONS),
        ]:
            c = ws.cell(row=row_num, column=3, value=val)
            c.number_format = fmt_code
            c.font = Fonts.body()
            c.border = Borders.bottom_only()
            c.alignment = Alignment(horizontal="right")

        # Enterprise & Equity Value
        fmt.apply_header_row(18, 2, 5, "ENTERPRISE VALUE → EQUITY VALUE BRIDGE")
        bridge_rows = [
            (19, "Sum of PV(FCFs)",         f"={pv_cols}",        NumFormats.USD_MILLIONS),
            (20, "PV of Terminal Value",    f"=C16",              NumFormats.USD_MILLIONS),
            (21, "Enterprise Value",        f"=C19+C20",          NumFormats.USD_MILLIONS),
            (22, "Less: Net Debt",          f"=-{asm}!C41",        NumFormats.USD_MILLIONS),
            (23, "Equity Value",            f"=C21+C22",          NumFormats.USD_MILLIONS),
            (24, "Shares Outstanding (M)",  f"={asm}!C40",         NumFormats.INTEGER),
            (25, "Implied Share Price",     f"=C23/C24",          NumFormats.USD_FULL),
        ]
        for row_num, label, formula, fmt_code in bridge_rows:
            fmt.apply_label_cell(row_num, 2, label)
            c = ws.cell(row=row_num, column=3, value=formula)
            c.number_format = fmt_code
            c.alignment = Alignment(horizontal="right")
            if row_num in [21, 23, 25]:
                c.font = Font(name="Calibri", size=12, bold=True, color=Colors.INST_BLUE)
                c.fill = PatternFill("solid", fgColor="EBF3FB")
                c.border = Borders.thick_bottom()
            else:
                c.font = Fonts.body()
                c.border = Borders.bottom_only()

        # Sensitivity: WACC vs Terminal Growth
        fmt.apply_header_row(27, 2, 9, "SENSITIVITY ANALYSIS — IMPLIED SHARE PRICE")
        ws.cell(row=28, column=2, value="WACC →").font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=28, column=2).alignment = Alignment(horizontal="right")

        wacc_vals = [
            float(self.a.get("wacc_sens_1", 0.08)),
            float(self.a.get("wacc_sens_2", 0.09)),
            float(self.a.get("wacc_sens_3", 0.10)),
            float(self.a.get("wacc_sens_4", 0.11)),
            float(self.a.get("wacc_sens_5", 0.12)),
        ]
        tg_vals = [0.015, 0.020, 0.025, 0.030, 0.035]

        ws.cell(row=27, column=2, value="Terminal Growth ↓ / WACC →")
        for j, w in enumerate(wacc_vals):
            c = ws.cell(row=28, column=3 + j, value=w)
            c.number_format = NumFormats.PERCENT_ONE
            c.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            c.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            c.alignment = Alignment(horizontal="center")

        for i, tg in enumerate(tg_vals):
            row = 29 + i
            tg_cell = ws.cell(row=row, column=2, value=tg)
            tg_cell.number_format = NumFormats.PERCENT_ONE
            tg_cell.font = Font(name="Calibri", size=10, bold=True, color=Colors.WHITE)
            tg_cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            tg_cell.alignment = Alignment(horizontal="center")

            ev_pv = "+".join([f"={fcf_sheet}!{get_column_letter(dc+k)}12/((1+{wacc_vals[0]})^{k+1})" for k in range(n)])
            for j, w in enumerate(wacc_vals):
                pv_fcfs = "+".join([
                    f"({fcf_sheet}!{get_column_letter(dc+k)}12/(1+{w})^{k+1})"
                    for k in range(n)
                ])
                tv = f"({fcf_sheet}!{last_fcf_col}12*(1+{tg})/({w}-{tg}))/(1+{w})^{n}"
                ev = f"={pv_fcfs}+{tv}"
                eq_val = f"({pv_fcfs}+{tv}-{asm}!C41)"
                share_price = f"=({pv_fcfs}+{tv}-{asm}!C41)/{asm}!C40"

                c = ws.cell(row=row, column=3 + j, value=share_price)
                c.number_format = NumFormats.USD_FULL
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center")

                # Color by relative value
                base_w = float(self.a.get("wacc_sens_3", 0.10))
                if w < base_w and tg > 0.02:
                    c.fill = Fills.sensitivity_high()
                elif w > base_w and tg < 0.025:
                    c.fill = Fills.sensitivity_low()
                else:
                    c.fill = Fills.sensitivity_mid()

        fmt.freeze_panes("A1")

    # ─── SHEET 7: DASHBOARD ─────────────────────────────────
    def _build_dashboard(self):
        ws, fmt = self.add_sheet("DASHBOARD", tab_color=Colors.ACCENT_GOLD)
        n = self.years

        # Layout
        ws.sheet_view.showGridLines = False
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 2
        for row in range(1, 60):
            ws.row_dimensions[row].height = 18

        # Dashboard fill
        for row in range(1, 60):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = Fills.dashboard()

        # Header bar
        for col in range(1, 20):
            ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)

        ws.merge_cells("C2:O2")
        title = ws["C2"]
        title.value = f"DCF MODEL DASHBOARD — {self.a.get('company_name','Company')}"
        title.font = Font(name="Calibri", size=16, bold=True, color=Colors.WHITE)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 30

        ws.merge_cells("C3:O3")
        sub = ws["C3"]
        sub.value = f"{self.a.get('currency','USD')} in Millions  |  {n}-Year Projection  |  Fiscal Year {self.base_year + 1}–{self.base_year + n}"
        sub.font = Font(name="Calibri", size=10, color="BDD7EE")
        sub.alignment = Alignment(horizontal="left", vertical="center")

        # KPI Cards — Row 5–8
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 22
        ws.row_dimensions[7].height = 38
        ws.row_dimensions[8].height = 8

        kpi_data = [
            ("C", "Enterprise Value ($M)",   "=VALUATION!C21", NumFormats.USD_MILLIONS),
            ("F", "Equity Value ($M)",        "=VALUATION!C23", NumFormats.USD_MILLIONS),
            ("I", "Implied Share Price",      "=VALUATION!C25", NumFormats.USD_FULL),
            ("L", "WACC",                     "=WACC!C22",      NumFormats.PERCENT_ONE),
            ("O", "Terminal Growth Rate",     "=ASSUMPTIONS!C38", NumFormats.PERCENT_ONE),
        ]

        for col_letter, label, formula, fmt_code in kpi_data:
            lbl_cell = ws[f"{col_letter}6"]
            lbl_cell.value = label
            lbl_cell.font = Fonts.kpi_label()
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.fill = Fills.kpi_card()
            lbl_cell.border = Border(
                top=Side(style="medium", color=Colors.INST_BLUE),
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE),
            )

            val_cell = ws[f"{col_letter}7"]
            val_cell.value = formula
            val_cell.font = Fonts.kpi_value()
            val_cell.number_format = fmt_code
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = Fills.kpi_card()
            val_cell.border = Border(
                left=Side(style="medium", color=Colors.INST_BLUE),
                right=Side(style="medium", color=Colors.INST_BLUE),
                bottom=Side(style="medium", color=Colors.INST_BLUE),
            )

        # Chart 1: Revenue & EBITDA Bar Chart
        rev_data = []
        ebitda_data = []
        years_list = [f"FY{self.base_year + i + 1}" for i in range(n)]

        for i in range(n):
            col = self.data_col_start + i
            cl = get_column_letter(col)
            rev_data.append(f"=INCOME_STATEMENT!{cl}7")
            ebitda_data.append(f"=INCOME_STATEMENT!{cl}15")

        # Write hidden data for charts
        chart_data_row = 50
        for i, yr in enumerate(years_list):
            ws.cell(row=chart_data_row, column=3 + i, value=yr)
            ws.cell(row=chart_data_row + 1, column=3 + i, value=rev_data[i])
            ws.cell(row=chart_data_row + 2, column=3 + i, value=ebitda_data[i])
            ws.cell(row=chart_data_row + 3, column=3 + i, value=f"=FCF_BRIDGE!{get_column_letter(self.data_col_start + i)}12")

        ws.cell(row=chart_data_row + 1, column=2, value="Revenue")
        ws.cell(row=chart_data_row + 2, column=2, value="EBITDA")
        ws.cell(row=chart_data_row + 3, column=2, value="UFCF")

        # Revenue chart
        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "clustered"
        chart1.title = "Revenue vs EBITDA ($M)"
        chart1.y_axis.title = "$ Millions"
        chart1.style = 10
        chart1.width = 14
        chart1.height = 10

        rev_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_data_row + 1, max_row=chart_data_row + 1)
        ebitda_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_data_row + 2, max_row=chart_data_row + 2)
        cats = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_data_row)

        chart1.add_data(rev_ref)
        chart1.add_data(ebitda_ref)
        chart1.set_categories(cats)
        from openpyxl.chart.series import SeriesLabel
        chart1.series[0].title = SeriesLabel(v="Revenue")
        chart1.series[1].title = SeriesLabel(v="EBITDA")
        ws.add_chart(chart1, "C10")

        # FCF Line chart
        chart2 = LineChart()
        chart2.title = "Free Cash Flow Trend ($M)"
        chart2.y_axis.title = "$ Millions"
        chart2.style = 10
        chart2.width = 14
        chart2.height = 10

        fcf_ref = Reference(ws, min_col=3, max_col=2 + n, min_row=chart_data_row + 3, max_row=chart_data_row + 3)
        chart2.add_data(fcf_ref)
        chart2.set_categories(cats)
        chart2.series[0].title = SeriesLabel(v="UFCF")
        ws.add_chart(chart2, "I10")

        # Secondary KPI row: EBITDA Margin, FCF Margin, Rev CAGR
        ws.row_dimensions[30].height = 22
        ws.row_dimensions[31].height = 38

        sec_kpis = [
            ("C", "EBITDA Margin (Exit Year)", f"=INCOME_STATEMENT!{get_column_letter(self.data_col_start + n - 1)}27", NumFormats.PERCENT_ONE),
            ("F", "Revenue CAGR",
             f"=(INCOME_STATEMENT!{get_column_letter(self.data_col_start+n-1)}7/INCOME_STATEMENT!{get_column_letter(self.data_col_start)}7)^(1/{n-1})-1",
             NumFormats.PERCENT_ONE),
            ("I", "FCF (Exit Year, $M)", f"=FCF_BRIDGE!{get_column_letter(self.data_col_start+n-1)}12", NumFormats.USD_MILLIONS),
            ("L", "EV / Exit EBITDA", f"=VALUATION!C21/INCOME_STATEMENT!{get_column_letter(self.data_col_start+n-1)}15", NumFormats.MULTIPLE),
            ("O", "Net Debt ($M)", "=ASSUMPTIONS!C41", NumFormats.USD_MILLIONS),
        ]

        for col_letter, label, formula, fmt_code in sec_kpis:
            lbl_cell = ws[f"{col_letter}30"]
            lbl_cell.value = label
            lbl_cell.font = Fonts.kpi_label()
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.fill = Fills.kpi_card()
            lbl_cell.border = Border(
                top=Side(style="medium", color=Colors.ACCENT_GOLD),
                left=Side(style="medium", color=Colors.ACCENT_GOLD),
                right=Side(style="medium", color=Colors.ACCENT_GOLD),
            )

            val_cell = ws[f"{col_letter}31"]
            val_cell.value = formula
            val_cell.font = Font(name="Calibri", size=18, bold=True, color=Colors.DARK_NAVY)
            val_cell.number_format = fmt_code
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = Fills.kpi_card()
            val_cell.border = Border(
                left=Side(style="medium", color=Colors.ACCENT_GOLD),
                right=Side(style="medium", color=Colors.ACCENT_GOLD),
                bottom=Side(style="medium", color=Colors.ACCENT_GOLD),
            )

        fmt.hide_gridlines()
