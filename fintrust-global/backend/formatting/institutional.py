"""
Institutional Formatting Engine
Financial Modeling Agent — Formatting Standards
All models must comply with these constants. No exceptions.
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ─────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────
class Colors:
    # Font Colors
    INPUT_BLUE       = "1F4E79"   # User inputs — blue
    FORMULA_BLACK    = "000000"   # Calculated cells — black
    EXTERNAL_GREEN   = "375623"   # External links — green
    NEGATIVE_RED     = "C00000"   # Negative values / errors
    POSITIVE_GREEN   = "375623"   # Positive variance

    # Background Colors
    HEADER_BG        = "D6DCE4"   # Section header background — grey
    SUBHEADER_BG     = "EEF2F7"   # Sub-section background — light grey
    COVER_BG         = "1F2D40"   # Cover page — dark navy
    COVER_ACCENT     = "2E75B6"   # Cover accent — institutional blue
    INPUT_BG         = "EBF3FB"   # Input cell background — light blue
    DASHBOARD_BG     = "F7F9FC"   # Dashboard background
    KPI_CARD_BG      = "FFFFFF"   # KPI card — white
    KPI_CARD_BORDER  = "2E75B6"   # KPI card border — blue
    SENSITIVITY_HIGH = "C6EFCE"   # Sensitivity high — green
    SENSITIVITY_LOW  = "FFC7CE"   # Sensitivity low — red
    SENSITIVITY_MID  = "FFEB9C"   # Sensitivity mid — yellow
    WHITE            = "FFFFFF"
    LIGHT_GREY       = "F2F2F2"
    MEDIUM_GREY      = "D6DCE4"
    DARK_NAVY        = "1F2D40"
    INST_BLUE        = "2E75B6"
    ACCENT_GOLD      = "C9A84C"

# ─────────────────────────────────────────────
# FONT STANDARDS
# ─────────────────────────────────────────────
class Fonts:
    FONT_NAME = "Calibri"
    SIZE_BODY       = 11
    SIZE_SUBHEADER  = 11
    SIZE_HEADER     = 12
    SIZE_TITLE      = 16
    SIZE_COVER_MAIN = 28
    SIZE_COVER_SUB  = 14
    SIZE_KPI_VALUE  = 20
    SIZE_KPI_LABEL  = 10

    @staticmethod
    def body():
        return Font(name="Calibri", size=11, color=Colors.FORMULA_BLACK)

    @staticmethod
    def input():
        return Font(name="Calibri", size=11, color=Colors.INPUT_BLUE)

    @staticmethod
    def header():
        return Font(name="Calibri", size=12, bold=True, color=Colors.FORMULA_BLACK)

    @staticmethod
    def section_header():
        return Font(name="Calibri", size=11, bold=True, color=Colors.FORMULA_BLACK)

    @staticmethod
    def title():
        return Font(name="Calibri", size=16, bold=True, color=Colors.DARK_NAVY)

    @staticmethod
    def cover_main():
        return Font(name="Calibri", size=28, bold=True, color=Colors.WHITE)

    @staticmethod
    def cover_sub():
        return Font(name="Calibri", size=14, color="BDD7EE")

    @staticmethod
    def kpi_value():
        return Font(name="Calibri", size=20, bold=True, color=Colors.INST_BLUE)

    @staticmethod
    def kpi_label():
        return Font(name="Calibri", size=10, color="595959")

    @staticmethod
    def negative():
        return Font(name="Calibri", size=11, color=Colors.NEGATIVE_RED)

    @staticmethod
    def external():
        return Font(name="Calibri", size=11, color=Colors.EXTERNAL_GREEN)

# ─────────────────────────────────────────────
# FILLS
# ─────────────────────────────────────────────
class Fills:
    @staticmethod
    def header_grey():
        return PatternFill("solid", fgColor=Colors.HEADER_BG)

    @staticmethod
    def subheader():
        return PatternFill("solid", fgColor=Colors.SUBHEADER_BG)

    @staticmethod
    def input_blue():
        return PatternFill("solid", fgColor=Colors.INPUT_BG)

    @staticmethod
    def cover_navy():
        return PatternFill("solid", fgColor=Colors.COVER_BG)

    @staticmethod
    def dashboard():
        return PatternFill("solid", fgColor=Colors.DASHBOARD_BG)

    @staticmethod
    def kpi_card():
        return PatternFill("solid", fgColor=Colors.KPI_CARD_BG)

    @staticmethod
    def sensitivity_high():
        return PatternFill("solid", fgColor=Colors.SENSITIVITY_HIGH)

    @staticmethod
    def sensitivity_low():
        return PatternFill("solid", fgColor=Colors.SENSITIVITY_LOW)

    @staticmethod
    def sensitivity_mid():
        return PatternFill("solid", fgColor=Colors.SENSITIVITY_MID)

    @staticmethod
    def white():
        return PatternFill("solid", fgColor=Colors.WHITE)

    @staticmethod
    def light_grey():
        return PatternFill("solid", fgColor=Colors.LIGHT_GREY)

# ─────────────────────────────────────────────
# BORDERS
# ─────────────────────────────────────────────
class Borders:
    @staticmethod
    def thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def bottom_only():
        return Border(bottom=Side(style="thin", color="BFBFBF"))

    @staticmethod
    def top_only():
        return Border(top=Side(style="thin", color="BFBFBF"))

    @staticmethod
    def thick_bottom():
        return Border(bottom=Side(style="medium", color=Colors.INST_BLUE))

    @staticmethod
    def kpi_card():
        s = Side(style="medium", color=Colors.KPI_CARD_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def none():
        return Border()

# ─────────────────────────────────────────────
# NUMBER FORMATS
# ─────────────────────────────────────────────
class NumFormats:
    USD_MILLIONS    = '$#,##0.0'
    USD_THOUSANDS   = '$#,##0'
    USD_FULL        = '$#,##0.00'
    PERCENT_ONE     = '0.0%'
    PERCENT_TWO     = '0.00%'
    MULTIPLE        = '0.0x'
    INTEGER         = '#,##0'
    DECIMAL_TWO     = '#,##0.00'
    YEAR            = '0'
    RATIO           = '0.0x'

# ─────────────────────────────────────────────
# COLUMN WIDTHS
# ─────────────────────────────────────────────
class ColWidths:
    LABEL           = 36    # Row description column
    YEAR            = 14    # Projection year columns
    SMALL           = 8     # Small spacer
    MEDIUM          = 18    # Medium data column
    LARGE           = 24    # Large label/description
    KPI_CARD        = 20    # Dashboard KPI card column

# ─────────────────────────────────────────────
# FORMATTING APPLIER CLASS
# ─────────────────────────────────────────────
class Formatter:

    def __init__(self, ws):
        self.ws = ws

    def apply_header_row(self, row, col_start, col_end, title):
        """Apply section header formatting across a row range"""
        cell = self.ws.cell(row=row, column=col_start, value=title)
        cell.font = Fonts.section_header()
        cell.fill = Fills.header_grey()
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = Borders.thick_bottom()

        for col in range(col_start + 1, col_end + 1):
            c = self.ws.cell(row=row, column=col)
            c.fill = Fills.header_grey()
            c.border = Borders.thick_bottom()

    def apply_input_cell(self, row, col, value=None, num_format=None):
        """Style a user-input cell — blue font, light blue fill"""
        cell = self.ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        cell.font = Fonts.input()
        cell.fill = Fills.input_blue()
        cell.border = Borders.thin()
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if num_format:
            cell.number_format = num_format
        return cell

    def apply_formula_cell(self, row, col, formula=None, num_format=None):
        """Style a formula cell — black font, white fill"""
        cell = self.ws.cell(row=row, column=col)
        if formula is not None:
            cell.value = formula
        cell.font = Fonts.body()
        cell.fill = Fills.white()
        cell.border = Borders.bottom_only()
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if num_format:
            cell.number_format = num_format
        return cell

    def apply_label_cell(self, row, col, value, indent=0):
        """Style a row label cell"""
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.font = Fonts.body()
        cell.fill = Fills.white()
        cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=indent
        )
        return cell

    def apply_total_row(self, row, col_start, col_end, label, num_format=None):
        """Apply total row formatting — bold with top/bottom border"""
        label_cell = self.ws.cell(row=row, column=col_start, value=label)
        label_cell.font = Font(name="Calibri", size=11, bold=True)
        label_cell.fill = Fills.subheader()
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        label_cell.border = Border(
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="double", color="000000")
        )
        for col in range(col_start + 1, col_end + 1):
            c = self.ws.cell(row=row, column=col)
            c.font = Font(name="Calibri", size=11, bold=True)
            c.fill = Fills.subheader()
            c.border = Border(
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="double", color="000000")
            )
            c.alignment = Alignment(horizontal="right", vertical="center")
            if num_format:
                c.number_format = num_format

    def set_column_widths(self, width_map):
        """Set column widths. width_map = {col_letter: width}"""
        for col_letter, width in width_map.items():
            self.ws.column_dimensions[col_letter].width = width

    def freeze_panes(self, cell_ref="B6"):
        """Apply institutional freeze panes"""
        self.ws.freeze_panes = cell_ref

    def hide_gridlines(self):
        """Remove gridlines — institutional standard"""
        self.ws.sheet_view.showGridLines = False

    def set_zoom(self, zoom=85):
        """Set sheet zoom level"""
        self.ws.sheet_view.zoomScale = zoom

    def apply_timeline_header(self, row, col_start, years, base_year=2024):
        """Write year headers across projection columns"""
        for i, yr in enumerate(range(base_year, base_year + years)):
            col = col_start + i
            cell = self.ws.cell(row=row, column=col, value=f"FY{yr}")
            cell.font = Font(name="Calibri", size=11, bold=True, color=Colors.WHITE)
            cell.fill = PatternFill("solid", fgColor=Colors.DARK_NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Borders.thin()

    def apply_units_label(self, row, col, unit_text="$ in Millions"):
        """Write units label — institutional requirement"""
        cell = self.ws.cell(row=row, column=col, value=unit_text)
        cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
        cell.alignment = Alignment(horizontal="left")

    def apply_kpi_card(self, row, col, label, value, num_format=None):
        """Create a KPI card block (3 rows x 2 cols)"""
        # Label row
        lbl = self.ws.cell(row=row, column=col, value=label)
        lbl.font = Fonts.kpi_label()
        lbl.fill = Fills.kpi_card()
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border = Border(
            top=Side(style="medium", color=Colors.KPI_CARD_BORDER),
            left=Side(style="medium", color=Colors.KPI_CARD_BORDER),
            right=Side(style="medium", color=Colors.KPI_CARD_BORDER)
        )

        # Value row
        val = self.ws.cell(row=row + 1, column=col, value=value)
        val.font = Fonts.kpi_value()
        val.fill = Fills.kpi_card()
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = Border(
            left=Side(style="medium", color=Colors.KPI_CARD_BORDER),
            right=Side(style="medium", color=Colors.KPI_CARD_BORDER),
            bottom=Side(style="medium", color=Colors.KPI_CARD_BORDER)
        )
        if num_format:
            val.number_format = num_format

        self.ws.row_dimensions[row].height = 20
        self.ws.row_dimensions[row + 1].height = 36

    def apply_sheet_title(self, row, col, title, subtitle=None):
        """Write sheet title block"""
        t = self.ws.cell(row=row, column=col, value=title)
        t.font = Fonts.title()
        t.alignment = Alignment(horizontal="left", vertical="center")
        if subtitle:
            s = self.ws.cell(row=row + 1, column=col, value=subtitle)
            s.font = Font(name="Calibri", size=10, italic=True, color="595959")
            s.alignment = Alignment(horizontal="left")
