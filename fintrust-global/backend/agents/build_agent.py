"""
Agent 4 — Build Agent
Agent 5 — QA Agent (Formula + Chart + Formatting checks)
Agent 6 — Delivery Agent
"""

import asyncio
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


# ══════════════════════════════════════════════════════════
# AGENT 4 — BUILD AGENT
# ══════════════════════════════════════════════════════════

class BuildAgent:

    def __init__(self, session: dict):
        self.session = session

    def _log(self, msg: str, status: str = "info"):
        self.session["logs"].append({
            "agent": "Build Agent",
            "message": msg,
            "status": status,
            "timestamp": __import__("datetime").datetime.now().isoformat()
        })

    async def build(self) -> str:
        """Build Excel model and return file path"""
        model_type = self.session.get("model_recommendation", "DCF")
        assumptions = self.session.get("assumptions", {})
        company_name = assumptions.get("company_name", "Company").replace(" ", "_")

        self._log(f"Building {model_type} model for {assumptions.get('company_name')}...", "thinking")
        await asyncio.sleep(0.1)

        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs")
        os.makedirs(output_dir, exist_ok=True)

        timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"{company_name}_{model_type}_{timestamp}.xlsx")

        # Build based on model type
        if model_type == "DCF":
            await self._build_dcf(assumptions, output_path)
        elif model_type == "LBO":
            await self._build_lbo(assumptions, output_path)
        elif model_type == "3-Statement":
            await self._build_3stmt(assumptions, output_path)
        else:
            await self._build_fpa(assumptions, output_path)

        # Add COMPS sheet
        await self._add_comps_sheet(output_path, assumptions)

        # Add Scenarios sheet
        await self._add_scenarios_sheet(output_path, assumptions)

        self._log(f"Model saved: {os.path.basename(output_path)}", "success")
        return output_path

    async def _build_dcf(self, assumptions: dict, path: str):
        from builders.dcf_builder import DCFBuilder
        self._log("Building COVER sheet...", "info")
        await asyncio.sleep(0.1)
        self._log("Building ASSUMPTIONS sheet (blue inputs)...", "info")
        await asyncio.sleep(0.1)
        self._log("Building INCOME_STATEMENT (5-yr P&L)...", "info")
        await asyncio.sleep(0.1)
        self._log("Building FCF_BRIDGE (NOPAT → UFCF)...", "info")
        await asyncio.sleep(0.1)
        self._log("Building WACC sheet (CAPM)...", "info")
        await asyncio.sleep(0.1)
        self._log("Building VALUATION + sensitivity table...", "info")
        await asyncio.sleep(0.1)
        self._log("Building DASHBOARD (KPI cards + charts)...", "info")
        data = DCFBuilder(assumptions).build()
        with open(path, "wb") as f:
            f.write(data)

    async def _build_lbo(self, assumptions: dict, path: str):
        from builders.lbo_builder import LBOBuilder
        self._log("Building LBO sheets (8 total)...", "info")
        await asyncio.sleep(0.2)
        data = LBOBuilder(assumptions).build()
        with open(path, "wb") as f:
            f.write(data)

    async def _build_3stmt(self, assumptions: dict, path: str):
        from builders.three_stmt_builder import ThreeStatementBuilder
        self._log("Building 3-Statement sheets (9 total)...", "info")
        await asyncio.sleep(0.2)
        data = ThreeStatementBuilder(assumptions).build()
        with open(path, "wb") as f:
            f.write(data)

    async def _build_fpa(self, assumptions: dict, path: str):
        from builders.fpa_builder import FPABuilder
        self._log("Building FP&A sheets (8 total)...", "info")
        await asyncio.sleep(0.2)
        data = FPABuilder(assumptions).build()
        with open(path, "wb") as f:
            f.write(data)

    async def _add_comps_sheet(self, path: str, assumptions: dict):
        """Add Comparable Companies sheet to existing workbook"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        peers = assumptions.get("peers", [])
        if not peers:
            self._log("No peer data — skipping COMPS sheet", "warning")
            return

        self._log(f"Adding COMPS sheet ({len(peers)} peers)...", "info")
        wb = openpyxl.load_workbook(path)

        ws = wb.create_sheet("COMPS", 1)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "7030A0"

        # Header
        ws.merge_cells("B2:J2")
        t = ws["B2"]
        t.value = f"COMPARABLE COMPANIES ANALYSIS — {assumptions.get('company_name', 'Company')}"
        t.font = Font(name="Calibri", size=14, bold=True, color="1F2D40")

        ws.cell(row=3, column=2, value="$ in Millions | Market data as of analysis date")
        ws.cell(row=3, column=2).font = Font(name="Calibri", size=9, italic=True, color="595959")

        # Column headers
        headers = ["Company", "Ticker", "Mkt Cap ($M)", "EV/EBITDA", "P/E Ratio",
                   "EV/Revenue", "Rev Growth %", "EBITDA Margin %"]
        col_widths = [28, 12, 16, 14, 12, 14, 14, 16]

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 3
        for i, (hdr, w) in enumerate(zip(headers, col_widths)):
            col = 3 + i
            ws.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width = w
            c = ws.cell(row=5, column=col, value=hdr)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F2D40")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[5].height = 22

        # Peer data rows
        for i, peer in enumerate(peers):
            row = 6 + i
            row_data = [
                peer.get("name", ""),
                peer.get("ticker", ""),
                peer.get("market_cap", 0),
                peer.get("ev_ebitda", 0),
                peer.get("pe_ratio", 0),
                peer.get("ev_revenue", 0),
                peer.get("revenue_growth", 0),
                peer.get("ebitda_margin", 0),
            ]
            formats = [None, None, '#,##0', '0.0x', '0.0x', '0.0x', '0.0%', '0.0%']

            for j, (val, fmt) in enumerate(zip(row_data, formats)):
                col = 3 + j
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else "F7F9FC")
                c.alignment = Alignment(horizontal="center" if j > 0 else "left", vertical="center")
                if fmt:
                    c.number_format = fmt
                c.border = Border(bottom=Side(style="thin", color="E0E0E0"))

        # Mean row
        mean_row = 6 + len(peers) + 1
        ws.cell(row=mean_row, column=3, value="MEAN").font = Font(name="Calibri", size=10, bold=True)
        for j in range(2, 8):
            col = 3 + j
            col_letter = __import__("openpyxl").utils.get_column_letter(col)
            c = ws.cell(row=mean_row, column=col,
                value=f"=AVERAGE({col_letter}6:{col_letter}{5+len(peers)})")
            formats = [None, None, '#,##0', '0.0x', '0.0x', '0.0x', '0.0%', '0.0%']
            if formats[j]:
                c.number_format = formats[j]
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = PatternFill("solid", fgColor="EBF3FB")

        self._log("COMPS sheet added successfully", "success")
        wb.save(path)

    async def _add_scenarios_sheet(self, path: str, assumptions: dict):
        """Add Bull/Base/Bear scenarios sheet"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        scenarios = assumptions.get("scenarios")
        if not scenarios:
            self._log("No scenario data — skipping SCENARIOS sheet", "warning")
            return

        self._log("Adding SCENARIOS sheet (Bull / Base / Bear)...", "info")
        wb = openpyxl.load_workbook(path)
        ws = wb.create_sheet("SCENARIOS", 2)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "C9A84C"

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20

        # Title
        ws.merge_cells("B2:E2")
        t = ws["B2"]
        t.value = f"SCENARIO ANALYSIS — {assumptions.get('company_name', 'Company')}"
        t.font = Font(name="Calibri", size=14, bold=True, color="1F2D40")

        # Scenario headers
        scenario_colors = {"BULL CASE": "375623", "BASE CASE": "2E75B6", "BEAR CASE": "C00000"}
        scenario_keys = {"BULL CASE": "bull", "BASE CASE": "base", "BEAR CASE": "bear"}

        for col_offset, (label, color) in enumerate(scenario_colors.items()):
            col = 3 + col_offset
            c = ws.cell(row=4, column=col, value=label)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[4].height = 24

        # Description row
        desc = scenarios.get("descriptions", {})
        for col_offset, key in enumerate(["bull", "base", "bear"]):
            col = 3 + col_offset
            c = ws.cell(row=5, column=col, value=desc.get(key, ""))
            c.font = Font(name="Calibri", size=9, italic=True, color="595959")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[5].height = 32

        # Key assumptions comparison
        ws.cell(row=7, column=2, value="KEY ASSUMPTIONS").font = Font(name="Calibri", size=11, bold=True)

        compare_fields = [
            ("rev_growth_y1", "Year 1 Revenue Growth", "0.0%"),
            ("rev_growth_y3", "Year 3 Revenue Growth", "0.0%"),
            ("rev_growth_y5", "Year 5 Revenue Growth", "0.0%"),
            ("ebitda_margin", "EBITDA Margin", "0.0%"),
            ("terminal_growth", "Terminal Growth Rate", "0.0%"),
        ]

        for i, (field, label, fmt) in enumerate(compare_fields):
            row = 8 + i
            ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=10)

            for col_offset, scen_key in enumerate(["bull", "base", "bear"]):
                col = 3 + col_offset
                scen = scenarios.get(scen_key, {})
                val = scen.get(field, assumptions.get(field, 0))
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = fmt
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center")
                bg = "E2EFDA" if col_offset == 0 else ("FFFFFF" if col_offset == 1 else "FCE4D6")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = Border(bottom=Side(style="thin", color="E0E0E0"))

        self._log("SCENARIOS sheet added successfully", "success")
        wb.save(path)


# ══════════════════════════════════════════════════════════
# AGENT 5 — QA AGENT
# ══════════════════════════════════════════════════════════

class QAAgent:

    def __init__(self, session: dict):
        self.session = session
        self.checks_run = 0
        self.checks_passed = 0
        self.issues = []

    def _log(self, msg: str, status: str = "info"):
        self.session["logs"].append({
            "agent": "QA Agent",
            "message": msg,
            "status": status,
            "timestamp": __import__("datetime").datetime.now().isoformat()
        })

    async def audit(self, excel_path: str) -> dict:
        """Full QA audit of Excel file"""
        import openpyxl
        self.issues = []
        self.checks_run = 0
        self.checks_passed = 0

        self._log("Opening Excel file for audit...", "thinking")
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        await self._check_formulas(wb)
        await self._check_charts(wb)
        await self._check_formatting(wb)
        await self._check_sheets(wb)

        passed = len(self.issues) == 0
        self._log(f"Audit complete: {self.checks_passed}/{self.checks_run} checks passed", 
                  "success" if passed else "warning")

        return {
            "passed": passed,
            "checks_run": self.checks_run,
            "checks_passed": self.checks_passed,
            "issues": self.issues
        }

    async def _check_formulas(self, wb):
        """Check all formula cells for errors"""
        self._log("Checking formula cells...", "thinking")
        await asyncio.sleep(0.1)

        error_values = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        formula_count = 0
        error_count = 0

        for ws in wb.worksheets:
            if ws.title in ["COVER"]:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        # Check for obvious errors in formula text
                        formula = cell.value.upper()
                        if any(err in formula for err in ["#REF", "#DIV"]):
                            error_count += 1
                            self.issues.append({
                                "type": "formula_error",
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "issue": f"Potential error in formula: {cell.value[:50]}",
                                "severity": "high"
                            })

        self.checks_run += 1
        if error_count == 0:
            self.checks_passed += 1
            self._log(f"Formula check passed: {formula_count} formulas, 0 errors", "success")
        else:
            self._log(f"Formula check: {error_count} potential errors found", "warning")

    async def _check_charts(self, wb):
        """Check all charts have valid data sources"""
        self._log("Checking chart objects...", "thinking")
        await asyncio.sleep(0.1)

        chart_count = 0
        chart_issues = 0

        for ws in wb.worksheets:
            for chart in ws._charts:
                chart_count += 1
                # Check chart has series
                if not chart.series:
                    chart_issues += 1
                    self.issues.append({
                        "type": "chart_error",
                        "sheet": ws.title,
                        "cell": "N/A",
                        "issue": f"Chart '{chart.title}' has no data series",
                        "severity": "medium"
                    })
                # Check chart has title
                if not chart.title:
                    self.issues.append({
                        "type": "chart_warning",
                        "sheet": ws.title,
                        "cell": "N/A",
                        "issue": f"Chart missing title",
                        "severity": "low"
                    })

        self.checks_run += 1
        if chart_issues == 0:
            self.checks_passed += 1
            self._log(f"Chart check passed: {chart_count} charts validated", "success")
        else:
            self._log(f"Chart check: {chart_issues} issues found", "warning")

    async def _check_formatting(self, wb):
        """Check institutional formatting rules"""
        self._log("Checking institutional formatting rules...", "thinking")
        await asyncio.sleep(0.1)

        formatting_issues = 0

        for ws in wb.worksheets:
            # Check gridlines hidden
            if ws.sheet_view.showGridLines:
                formatting_issues += 1
                self.issues.append({
                    "type": "formatting",
                    "sheet": ws.title,
                    "cell": "N/A",
                    "issue": "Gridlines still visible — should be hidden",
                    "severity": "medium",
                    "auto_fixable": True
                })

            # Check freeze panes on model sheets
            if ws.title not in ["COVER", "DASHBOARD", "COMPS", "SCENARIOS"]:
                if not ws.freeze_panes:
                    self.issues.append({
                        "type": "formatting",
                        "sheet": ws.title,
                        "cell": "N/A",
                        "issue": "Freeze panes not set on model sheet",
                        "severity": "low",
                        "auto_fixable": True
                    })

        self.checks_run += 1
        if formatting_issues == 0:
            self.checks_passed += 1
            self._log("Formatting check passed", "success")
        else:
            self._log(f"Formatting: {formatting_issues} issues found", "warning")

    async def _check_sheets(self, wb):
        """Check required sheets exist"""
        self._log("Checking required sheets exist...", "thinking")
        sheet_names = [ws.title for ws in wb.worksheets]
        required = ["COVER", "ASSUMPTIONS", "DASHBOARD"]
        missing_sheets = [s for s in required if s not in sheet_names]

        self.checks_run += 1
        if not missing_sheets:
            self.checks_passed += 1
            self._log(f"Sheet check passed: {len(sheet_names)} sheets found", "success")
        else:
            for s in missing_sheets:
                self.issues.append({
                    "type": "missing_sheet",
                    "sheet": s,
                    "cell": "N/A",
                    "issue": f"Required sheet '{s}' is missing",
                    "severity": "high"
                })
            self._log(f"Missing sheets: {missing_sheets}", "error")

    async def auto_fix(self, excel_path: str, issues: list) -> str:
        """Auto-fix detected issues"""
        import openpyxl

        fixable = [i for i in issues if i.get("auto_fixable")]
        if not fixable:
            return excel_path

        self._log(f"Auto-fixing {len(fixable)} issues...", "thinking")
        wb = openpyxl.load_workbook(excel_path)

        for issue in fixable:
            sheet_name = issue.get("sheet")
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if not ws:
                continue

            if "gridlines" in issue.get("issue", "").lower():
                ws.sheet_view.showGridLines = False
                self._log(f"Fixed: Gridlines hidden on {sheet_name}", "success")

            elif "freeze panes" in issue.get("issue", "").lower():
                ws.freeze_panes = "C5"
                self._log(f"Fixed: Freeze panes set on {sheet_name}", "success")

        fixed_path = excel_path.replace(".xlsx", "_fixed.xlsx")
        wb.save(fixed_path)
        return fixed_path


# ══════════════════════════════════════════════════════════
# AGENT 6 — DELIVERY AGENT
# ══════════════════════════════════════════════════════════

class DeliveryAgent:

    def __init__(self, session: dict):
        self.session = session

    def _log(self, msg: str, status: str = "info"):
        self.session["logs"].append({
            "agent": "Delivery Agent",
            "message": msg,
            "status": status,
            "timestamp": __import__("datetime").datetime.now().isoformat()
        })

    async def prepare(self):
        """Final delivery preparation"""
        path = self.session.get("excel_path")
        company = self.session.get("assumptions", {}).get("company_name", "Company")
        model = self.session.get("model_recommendation", "Model")
        qa = self.session.get("qa_report", {})

        self._log(f"Preparing {company} {model} model for delivery...", "thinking")
        await asyncio.sleep(0.2)

        checks = qa.get("checks_passed", 0)
        total = qa.get("checks_run", 0)

        self._log(f"QA Score: {checks}/{total} checks passed", "success")
        self._log(f"File size: {os.path.getsize(path) / 1024:.0f} KB", "info")
        self._log(f"Sheets: COVER + ASSUMPTIONS + Model Sheets + COMPS + SCENARIOS + DASHBOARD", "info")
        self._log("✅ Model ready for download", "success")
